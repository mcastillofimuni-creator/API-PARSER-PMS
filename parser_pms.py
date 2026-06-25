import re
import warnings
from datetime import datetime, date, time
from openpyxl.utils import get_column_letter

import openpyxl
import pandas as pd
from rapidfuzz import fuzz

warnings.filterwarnings("ignore")


# ============================================================
# CONFIGURACIÓN GENERAL
# ============================================================

HOJAS_EXCLUIR = [
    "FORMULARIO",
    "CLEAN",
    "RESUMEN",
    "INSTRUCCIONES",
    "BD",
    "BASE",
    "LISTA",
    "CONFIG",
    "DICCIONARIO",
    "TABLA",
    "MAESTRA",
    "DISTRIBUCION",
    "DISTRIBUCIÓN",
    "DISTRIBUTION",
]

CAMPOS = {
    "ot_grafo": [
        "n ot", "nro ot", "ot", "grafo", "ot grafo",
        "n°ot / grafo", "n°ot", "n° ot", "nº ot",
        "n°ot/grafo", "n ot grafo", "orden de trabajo"
    ],
    "central": ["central", "planta", "sede"],
    "unidad": ["grupo", "unidad", "grupo unidad", "unidad generadora"],
    "sistema": ["sistema", "sist", "sistema principal"],
    "equipo": [
        "sub sistema/equipo",
        "sub sistema equipo",
        "sub sistema",
        "subsistema",
        "equipo",
        "componente",
        "activo",
    ],
    "cod_pm_aviso": ["cod pm / aviso", "cod pm", "aviso", "codigo pm", "código pm"],
    "pedido": ["n pedido", "pedido", "n° pedido", "nº pedido"],
    "motivo": [
        "motivo",
        "descripcion del trabajo",
        "descripción del trabajo",
        "trabajo",
        "tarea",
        "actividad",
        "descripcion",
        "descripción",
    ],
    "tipo_mant": ["tipo mant", "tipo mantenimiento", "tipo de mantenimiento"],
    "condicion": ["condicion", "condición"],
    "riesgo": ["riesgo critico", "riesgo crítico", "riesgo", "criticidad"],
    "area_solicitante": ["area solicitante", "área solicitante"],
    "inspector": ["inspector", "supervisor", "supervisor orygen", "responsable orygen"],
    "rt_terceros": [
        "rt terceros",
        "responsable tercero",
        "rt",
        "responsable contratista",
        "responsable proveedor",
    ],
    "recursos": ["recursos", "personal", "cantidad personal"],
    "hora_inicio": ["hora inicio", "h inicio", "inicio hora"],
    "hora_fin": ["hora fin", "h fin", "fin hora"],
    "dias": ["dias", "días", "dia", "día"],
    "fecha_inicio": ["fecha inicio", "f inicio", "inicio"],
    "fecha_fin": ["fecha fin", "f fin", "fin"],
    "proveedor": ["empresa", "proveedor", "contratista"],
    "codigo_actividad": ["codigo de actividad", "código de actividad"],
    "texto_explicativo": ["texto explicativo"],
    "riesgo_existente": ["riesgo existente"],
    "riesgo_introducido": ["riesgo introducido"],
    "observacion": ["observacion", "observación"],
}


# ============================================================
# DICCIONARIO INTELIGENTE OT VS AVISO
# ============================================================
# Estos pesos no bloquean por sí solos. Generan un score de sospecha
# para advertir cuando un número informado como OT podría ser un Aviso SAP.

AVISO_RANGO_100_MIN = 10000030
AVISO_RANGO_100_MAX = 10002533

PALABRAS_AVISO_ALTA = {
    "falla": 25,
    "fallo": 25,
    "inoperativo": 25,
    "alarma": 22,
    "fuga": 25,
    "rotura": 22,
    "perdida": 20,
    "bajo": 18,
    "baja": 18,
    "alto": 18,
    "alta": 18,
    "nivel": 16,
    "presion": 18,
    "temperatura": 18,
    "desgaste": 20,
    "filtracion": 20,
    "pase": 18,
    "trip": 20,
    "ruido": 18,
    "corrosion": 18,
    "senal": 18,
    "sensor": 14,
    "error": 18,
    "anormal": 16,
    "intermitente": 16,
    "comunicacion": 14,
    "averia": 22,
    "defecto": 18,
    "dañado": 18,
    "danado": 18,
}

FRASES_AVISO_ALTA = {
    "fuga aceite": 30,
    "fuga de aceite": 30,
    "fuga agua": 28,
    "fuga de agua": 28,
    "fuga aire": 26,
    "fuga de aire": 26,
    "baja presion": 28,
    "bajo nivel": 28,
    "alarma bajo": 26,
    "senal alarma": 25,
    "senal falla": 28,
    "falla sensor": 25,
    "desgaste interno": 24,
    "desgaste bomba": 24,
    "pase valvula": 22,
    "perdida comunicacion": 24,
    "no comunica": 22,
    "no responde": 22,
    "fuera de servicio": 18,
}

PALABRAS_PREVENTIVO = {
    "mantto": -22,
    "manto": -22,
    "mantenimiento": -18,
    "preventivo": -24,
    "prev": -20,
    "calibracion": -24,
    "instrumentacion": -18,
    "prueba": -12,
    "pruebas": -12,
    "anual": -16,
    "limpieza": -14,
    "inspeccion": -10,
    "analisis": -10,
    "aceite": -8,
    "reconfig": -18,
    "sustitucion": -14,
}


# ============================================================
# UTILIDADES
# ============================================================

def normalizar_texto(x):
    if x is None:
        return ""

    x = str(x).strip().lower()
    x = x.replace("\n", " ")
    x = re.sub(r"\s+", " ", x)

    reemplazos = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u",
        "Á": "a", "É": "e", "Í": "i", "Ó": "o", "Ú": "u",
        "°": "", "º": "", ".": "", ":": "", ";": "",
        "/": " ", "\\": " ", "-": " ", "_": " ",
    }

    for a, b in reemplazos.items():
        x = x.replace(a, b)

    x = re.sub(r"\s+", " ", x)
    return x.strip()


def limpiar_valor(x):
    if x is None:
        return ""

    if isinstance(x, float) and pd.isna(x):
        return ""

    if isinstance(x, datetime):
        return x.strftime("%Y-%m-%d %H:%M")

    if isinstance(x, date):
        return x.strftime("%Y-%m-%d")

    if isinstance(x, time):
        return x.strftime("%H:%M")

    x = str(x).upper().strip()
    x = x.replace("\n", " ")
    x = re.sub(r"\s+", " ", x)

    if x in ["NAN", "NONE", "NULL"]:
        return ""

    return x


def txt_upper(x):
    return limpiar_valor(x)


def esta_vacio(x):
    v = txt_upper(x)
    return v in [
        "",
        "NA",
        "N/A",
        "NONE",
        "NAN",
        "NULL",
        "-",
        "--",
        "S/I",
        "SIN INFO",
        "EMPTY",
    ]


def hoja_valida_para_programa(nombre_hoja):
    h = limpiar_valor(nombre_hoja)
    return not any(palabra in h for palabra in HOJAS_EXCLUIR)


def limpiar_ot(valor):
    if valor is None:
        return ""

    if isinstance(valor, float) and pd.isna(valor):
        return ""

    if isinstance(valor, int):
        return str(valor)

    if isinstance(valor, float):
        if valor.is_integer():
            return str(int(valor))
        return str(valor).strip()

    v = str(valor).strip()

    if re.fullmatch(r"\d+\.0", v):
        return v.split(".")[0]

    return re.sub(r"\D", "", v)


def es_numero_pedido(valor):
    """
    Detecta números de pedido SAP informados en el PMS.
    Patrones permitidos:
    - 3500xxxx
    - 4500xxxx

    Se tratan como PEDIDO, no como OT inválida.
    """
    n = limpiar_ot(valor)
    return bool(re.fullmatch(r"(3500|4500)\d{4,8}", n))


def extraer_numero_pedido(valor):
    n = limpiar_ot(valor)
    return n if es_numero_pedido(n) else ""


def actualizar_datos_originales_pedido_desde_ot(datos_originales, pedido_detectado):
    """
    Marca en datos_originales que el número de pedido fue detectado en OT/Grafo.
    No borra el dato original; deja trazabilidad para que el generador PMS pueda
    moverlo a la columna Pedido más adelante.
    """
    datos = json_seguro(datos_originales)

    if not isinstance(datos, dict):
        datos = {}

    datos["pedido_detectado_en_ot"] = True
    datos["numero_pedido_detectado"] = pedido_detectado

    campos_detectados = datos.get("campos_detectados")
    if not isinstance(campos_detectados, dict):
        campos_detectados = {}
        datos["campos_detectados"] = campos_detectados

    campos_detectados["pedido"] = pedido_detectado

    mapa_campos = datos.get("mapa_campos")
    columnas_excel = datos.get("columnas_excel")

    if isinstance(mapa_campos, dict) and isinstance(columnas_excel, dict):
        info_pedido = mapa_campos.get("pedido")
        if isinstance(info_pedido, dict):
            col_letra = info_pedido.get("col_letra")
            if col_letra:
                columnas_excel[col_letra] = pedido_detectado

    return datos


def convertir_a_str_seguro(valor):
    if valor is None:
        return ""

    if isinstance(valor, float) and pd.isna(valor):
        return ""

    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d %H:%M")

    if isinstance(valor, date):
        return valor.strftime("%Y-%m-%d")

    if isinstance(valor, time):
        return valor.strftime("%H:%M")

    valor = str(valor).strip()

    if valor.upper() in ["NAN", "NONE", "NULL"]:
        return ""

    return valor


def json_seguro(valor):
    """
    Convierte valores de Excel/Pandas/OpenPyXL en JSON seguro para Supabase.
    """
    if valor is None:
        return None

    if isinstance(valor, float) and pd.isna(valor):
        return None

    if isinstance(valor, (datetime, date, time)):
        return convertir_a_str_seguro(valor)

    if isinstance(valor, dict):
        return {str(k): json_seguro(v) for k, v in valor.items()}

    if isinstance(valor, list):
        return [json_seguro(v) for v in valor]

    if isinstance(valor, tuple):
        return [json_seguro(v) for v in valor]

    if isinstance(valor, (str, int, float, bool)):
        return valor

    return str(valor)


def normalizar_central_operativa(valor):
    """
    Normaliza nombres de central para comparar.
    """

    v = limpiar_valor(valor)

    if not v:
        return ""

    if "SANTA ROSA" in v:
        return "SANTA ROSA"

    if "VENTANILLA" in v:
        return "VENTANILLA"

    if "DISTRIBUCION" in v or "DISTRIBUCIÓN" in v:
        return "DISTRIBUCION"

    return v


def obtener_central_presentada_desde_archivo_info(archivo_info):
    if not isinstance(archivo_info, dict):
        return ""

    return (
        archivo_info.get("central_presentada")
        or archivo_info.get("central")
        or archivo_info.get("central_presentada_norm")
        or ""
    )


# ============================================================
# DETECCIÓN DE ENCABEZADOS
# ============================================================

def identificar_campo(texto, umbral=78):
    texto_norm = normalizar_texto(texto)

    if not texto_norm:
        return None, 0

    reglas_directas = {
        "sub sistema equipo": "equipo",
        "sub sistema/equipo": "equipo",
        "subsistema": "equipo",
        "sub sistema": "equipo",
        "equipo": "equipo",
        "sistema": "sistema",
        "sist": "sistema",
        "motivo": "motivo",
        "actividad": "motivo",
        "descripcion": "motivo",
        "descripcion del trabajo": "motivo",
        "descripcion trabajo": "motivo",
        "empresa": "proveedor",
        "proveedor": "proveedor",
        "contratista": "proveedor",
        "condicion": "condicion",
        "condición": "condicion",
        "riesgo critico": "riesgo",
        "riesgo crítico": "riesgo",
        "riesgo": "riesgo",
        "inspector": "inspector",
        "supervisor": "inspector",
        "rt terceros": "rt_terceros",
        "responsable tercero": "rt_terceros",
        "hora inicio": "hora_inicio",
        "hora fin": "hora_fin",
        "dias": "dias",
        "días": "dias",
        "n ot grafo": "ot_grafo",
        "n°ot grafo": "ot_grafo",
        "n° ot grafo": "ot_grafo",
        "ot grafo": "ot_grafo",
        "ot": "ot_grafo",
        "grafo": "ot_grafo",
        "codigo de actividad": "codigo_actividad",
        "código de actividad": "codigo_actividad",
        "texto explicativo": "texto_explicativo",
        "riesgo existente": "riesgo_existente",
        "riesgo introducido": "riesgo_introducido",
    }

    if texto_norm in reglas_directas:
        return reglas_directas[texto_norm], 100

    mejor_campo = None
    mejor_score = 0

    for campo, sinonimos in CAMPOS.items():
        for sinonimo in sinonimos:
            s_norm = normalizar_texto(sinonimo)

            score = max(
                fuzz.ratio(texto_norm, s_norm),
                fuzz.partial_ratio(texto_norm, s_norm),
            )

            if score > mejor_score:
                mejor_score = score
                mejor_campo = campo

    if mejor_score >= umbral:
        return mejor_campo, mejor_score

    return None, mejor_score


def detectar_encabezado_en_hoja(ws, max_filas=90, max_columnas=80, min_campos=5):
    candidatos = []

    max_row = min(ws.max_row or 1, max_filas)
    max_col = min(ws.max_column or 1, max_columnas)

    for fila_idx, fila in enumerate(
        ws.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        ),
        start=1,
    ):
        campos_detectados = {}

        for col_idx, valor in enumerate(fila, start=1):
            campo, score = identificar_campo(valor)

            if campo and campo not in campos_detectados:
                campos_detectados[campo] = {
                    "col": col_idx,
                    "texto": convertir_a_str_seguro(valor),
                    "score": score,
                }

        campos_clave = {
            "ot_grafo",
            "central",
            "unidad",
            "sistema",
            "equipo",
            "motivo",
            "condicion",
            "riesgo",
            "inspector",
            "rt_terceros",
        }

        bonus = len(campos_clave.intersection(campos_detectados.keys()))
        total = len(campos_detectados) + bonus

        if len(campos_detectados) >= min_campos:
            candidatos.append({
                "fila": fila_idx,
                "total": total,
                "campos": campos_detectados,
            })

    if not candidatos:
        return None

    return sorted(candidatos, key=lambda x: x["total"], reverse=True)[0]


# ============================================================
# EXTRACCIÓN DE ACTIVIDADES
# ============================================================

def fila_tiene_datos(row_dict):
    """
    Determina si la fila parece una actividad real.

    Ajuste importante:
    No basta con que exista una descripción en MOTIVO. Algunos formatos tienen
    bloques informativos como "ACTIVIDADES PROPUESTAS" o listas auxiliares
    donde solo se llena MOTIVO + columnas de costo/comentario. Esas filas no
    deben entrar al PMS ni a la validación SAP.
    """
    motivo = row_dict.get("motivo", "")

    if esta_vacio(motivo):
        return False

    campos_operativos_fuertes = [
        "ot_grafo",
        "central",
        "unidad",
        "sistema",
        "equipo",
        "inspector",
        "rt_terceros",
        "fecha_inicio",
        "fecha_fin",
        "hora_inicio",
        "hora_fin",
    ]

    fuertes_no_vacios = [
        c for c in campos_operativos_fuertes
        if not esta_vacio(row_dict.get(c, ""))
    ]

    # Actividad real: motivo + al menos dos datos operativos fuertes.
    # Evita filas de comentario/propuesta donde solo aparece la descripción.
    return len(fuertes_no_vacios) >= 2


def detectar_central_en_hoja(ws, nombre_hoja, max_filas=15, max_columnas=20):
    """
    Detecta la central real de la hoja leyendo el contenido superior.
    Esto evita confiar ciegamente en el nombre de la pestaña, porque algunos
    proveedores pueden nombrar una hoja como Ventanilla aunque el contenido sea
    de Santa Rosa, o viceversa.
    """
    textos = [limpiar_valor(nombre_hoja)]

    max_row = min(ws.max_row or 1, max_filas)
    max_col = min(ws.max_column or 1, max_columnas)

    for fila in ws.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
        values_only=True,
    ):
        for valor in fila:
            v = limpiar_valor(valor)
            if v:
                textos.append(v)

    texto = " ".join(textos)

    if "SANTA ROSA" in texto or "CTSR" in texto or "C.T. SANTA ROSA" in texto or "CCTT SANTA ROSA" in texto:
        return "SANTA ROSA"

    if "VENTANILLA" in texto or "CTVE" in texto or "C.C. VENTANILLA" in texto or "C.T. VENTANILLA" in texto:
        return "VENTANILLA"

    return ""


def construir_datos_originales(nombre_hoja, fila_excel, fila, campos, central_hoja_detectada=""):
    """
    Guarda la fila completa original para luego reconstruir el Excel consolidado.

    Incluye:
    - columnas_excel: valores por letra de columna A, B, C...
    - campos_detectados: valores según los encabezados detectados
    - mapa_campos: columna detectada para cada campo
    """
    columnas_excel = {}

    for idx, valor in enumerate(fila, start=1):
        letra = get_column_letter(idx)
        columnas_excel[letra] = json_seguro(valor)

    campos_detectados = {}

    for campo, info in campos.items():
        col_idx = info["col"]

        if col_idx <= len(fila):
            campos_detectados[campo] = json_seguro(fila[col_idx - 1])
        else:
            campos_detectados[campo] = None

    mapa_campos = {
        campo: {
            "col": info.get("col"),
            "col_letra": get_column_letter(info.get("col")),
            "texto_header": info.get("texto"),
            "score": info.get("score"),
        }
        for campo, info in campos.items()
    }

    return {
        "hoja": nombre_hoja,
        "central_hoja_detectada": central_hoja_detectada,
        "fila_excel": fila_excel,
        "columnas_excel": columnas_excel,
        "campos_detectados": campos_detectados,
        "mapa_campos": mapa_campos,
    }


def extraer_actividades(nombre_archivo):
    wb = openpyxl.load_workbook(
        nombre_archivo,
        data_only=True,
        read_only=True,
    )

    actividades = []
    hojas_revisadas = []

    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        estado_hoja = getattr(ws, "sheet_state", "visible")

        if estado_hoja != "visible":
            continue

        if not hoja_valida_para_programa(nombre_hoja):
            continue

        central_hoja_detectada = detectar_central_en_hoja(ws, nombre_hoja)

        det = detectar_encabezado_en_hoja(ws)

        if not det:
            continue

        fila_header = det["fila"]
        campos = det["campos"]

        hojas_revisadas.append({
            "hoja": nombre_hoja,
            "estado": estado_hoja,
            "central_hoja_detectada": central_hoja_detectada,
            "fila_encabezado": fila_header,
            "campos_detectados": ", ".join(campos.keys()),
        })

        max_filas_lectura = min(ws.max_row or fila_header, fila_header + 500)
        max_col_lectura = min(ws.max_column or 1, 80)

        for fila_idx, fila in enumerate(
            ws.iter_rows(
                min_row=fila_header + 1,
                max_row=max_filas_lectura,
                min_col=1,
                max_col=max_col_lectura,
                values_only=True,
            ),
            start=fila_header + 1,
        ):
            row_dict = {
                "hoja": nombre_hoja,
                "central_hoja_detectada": central_hoja_detectada,
                "fila_excel": fila_idx,
            }

            for campo, info in campos.items():
                col_idx = info["col"]

                if col_idx <= len(fila):
                    row_dict[campo] = fila[col_idx - 1]
                else:
                    row_dict[campo] = None

            row_dict["datos_originales"] = construir_datos_originales(
                nombre_hoja=nombre_hoja,
                fila_excel=fila_idx,
                fila=fila,
                campos=campos,
                central_hoja_detectada=central_hoja_detectada,
            )

            if fila_tiene_datos(row_dict):
                actividades.append(row_dict)

    wb.close()

    df = pd.DataFrame(actividades)
    df_hojas = pd.DataFrame(hojas_revisadas)

    if df.empty:
        return df, df_hojas

    columnas_ordenadas = [
        "hoja",
        "central_hoja_detectada",
        "fila_excel",
        "proveedor",
        "ot_grafo",
        "central",
        "unidad",
        "sistema",
        "equipo",
        "motivo",
        "tipo_mant",
        "condicion",
        "riesgo",
        "area_solicitante",
        "inspector",
        "rt_terceros",
        "recursos",
        "hora_inicio",
        "hora_fin",
        "dias",
        "fecha_inicio",
        "fecha_fin",
        "cod_pm_aviso",
        "pedido",
        "codigo_actividad",
        "texto_explicativo",
        "riesgo_existente",
        "riesgo_introducido",
        "observacion",
        "datos_originales",
    ]

    columnas_existentes = [c for c in columnas_ordenadas if c in df.columns]
    otras = [c for c in df.columns if c not in columnas_existentes]

    return df[columnas_existentes + otras], df_hojas


# ============================================================
# NORMALIZACIONES
# ============================================================

def normalizar_unidad(x):
    x = limpiar_valor(x)

    if not x:
        return ""

    m = re.search(r"\bTG\s*[-]?\s*(\d+)\b", x)
    if m:
        return f"TG{m.group(1)}"

    m = re.search(r"\bG\s*[-]?\s*(\d+)\b", x)
    if m:
        return f"G{m.group(1)}"

    return x


def normalizar_sistema(x):
    x = limpiar_valor(x)

    if not x:
        return ""

    if "TRAFO" in x or "TRANSFORM" in x:
        return "TRANSFORMADOR"

    if "INCENDIO" in x or "SCI" in x or "CONTRA INCENDIO" in x:
        return "SISTEMA CONTRA INCENDIO"

    if "PROTECCION" in x or "PROTECCIONES" in x or "RELÉ" in x or "RELE" in x:
        return "PROTECCIONES"

    if "EXCIT" in x or "AVR" in x:
        return "EXCITACION / AVR"

    if "UPS" in x or "BATER" in x or "CARGADOR" in x:
        return "SERVICIOS AUXILIARES DC/AC"

    if "AGUA" in x and "REFRIG" in x:
        return "AGUA DE REFRIGERACION"

    if "TURBINA" in x or "COMPRESOR" in x:
        return "TURBINA-COMPRESOR"

    if "GASODUCTO" in x:
        return "GASODUCTO"

    if "DIESEL" in x or "DIÉSEL" in x:
        return "DIESEL"

    return x


def inferir_activo_padre(row):
    unidad = row.get("unidad_norm", "")
    sistema = row.get("sistema_norm", "")
    equipo = limpiar_valor(row.get("equipo", ""))

    texto = " ".join([unidad, sistema, equipo])

    if "TRANSFORMADOR" in texto or "TRAFO" in texto:
        return f"TRANSFORMADOR {unidad}".strip()

    if "SISTEMA CONTRA INCENDIO" in texto or "SCI" in texto or "INCENDIO" in texto:
        if "TRAFO" in texto or "TRANSFORM" in texto:
            return f"TRANSFORMADOR {unidad}".strip()
        return f"SCI {unidad}".strip()

    if "PROTECCIONES" in texto:
        return f"PROTECCIONES {unidad}".strip()

    if sistema:
        return f"{sistema} {unidad}".strip()

    return unidad


def es_fila_actividad_real(row):
    proveedor = limpiar_valor(row.get("proveedor", ""))
    central = limpiar_valor(row.get("central", ""))
    unidad = limpiar_valor(row.get("unidad", ""))
    sistema = limpiar_valor(row.get("sistema", ""))
    equipo = limpiar_valor(row.get("equipo", ""))
    motivo = limpiar_valor(row.get("motivo", ""))
    condicion = limpiar_valor(row.get("condicion", ""))
    inspector = limpiar_valor(row.get("inspector", ""))
    rt = limpiar_valor(row.get("rt_terceros", ""))
    ot = limpiar_valor(row.get("ot_grafo", ""))
    fecha_inicio = limpiar_valor(row.get("fecha_inicio", ""))
    fecha_fin = limpiar_valor(row.get("fecha_fin", ""))
    hora_inicio = limpiar_valor(row.get("hora_inicio", ""))
    hora_fin = limpiar_valor(row.get("hora_fin", ""))

    texto_total = " ".join([
        proveedor, central, unidad, sistema, equipo, motivo, condicion,
        inspector, rt, ot, fecha_inicio, fecha_fin, hora_inicio, hora_fin,
    ])

    palabras_no_actividad = [
        "TOTAL", "SUBTOTAL", "LEYENDA", "OBSERVACION", "OBSERVACIONES",
        "SEMANA", "PROGRAMA SEMANAL", "NOTA", "FORMULARIO", "CLEAN SHEET",
        "APROBADO", "REVISADO", "ELABORADO", "ACTIVIDADES PROPUESTAS",
        "PENDIENTES / ADICIONALES",
    ]

    if any(p in texto_total for p in palabras_no_actividad):
        return False

    if motivo in ["", "NONE", "NAN", "NA", "NULL", "N/A", "-", "EMPTY"]:
        return False

    # Una actividad real debe tener una descripción y al menos dos datos
    # operativos fuertes. Esto evita que se lean bloques auxiliares/propuestos
    # que no pertenecen al programa semanal validable.
    campos_fuertes = [
        central, unidad, sistema, equipo, inspector, rt, ot,
        fecha_inicio, fecha_fin, hora_inicio, hora_fin,
    ]

    fuertes_llenos = [
        c for c in campos_fuertes
        if c not in ["", "NONE", "NAN", "NA", "NULL", "N/A", "-", "EMPTY"]
    ]

    return len(fuertes_llenos) >= 2


# ============================================================
# VALIDACIONES DE FORMA
# ============================================================

def validar_ot(valor):
    """
    OT válida:
    - 8 o 9 dígitos.
    - Debe iniciar con 100, 200, 300 o 400.

    Nota: los números 100 bajos pueden parecer Avisos SAP; eso se evalúa
    después con evaluar_ot_vs_aviso(), sin bloquear automáticamente.
    """

    v_limpio = limpiar_ot(valor)

    if v_limpio == "":
        return "VACIA", v_limpio

    if not re.fullmatch(r"\d+", v_limpio):
        return "INVALIDA", v_limpio

    if es_numero_pedido(v_limpio):
        return "PEDIDO_EN_CAMPO_OT", v_limpio

    if len(v_limpio) not in (8, 9):
        return "LONGITUD_INVALIDA", v_limpio

    if v_limpio[0] not in ["1", "2", "3", "4"]:
        return "PRIMER_DIGITO_INVALIDO", v_limpio

    if v_limpio[1:3] != "00":
        return "PATRON_INVALIDO", v_limpio

    if re.fullmatch(r"[1234]00\d{5,6}", v_limpio):
        return "VALIDA", v_limpio

    return "INVALIDA", v_limpio


def score_texto_aviso(texto):
    """
    Evalúa si la descripción parece correctiva/aviso o preventiva.
    Devuelve score y motivos. Score positivo = más sospecha de Aviso.
    """
    texto_norm = normalizar_texto(texto)
    score = 0
    motivos = []

    if not texto_norm:
        return 0, motivos

    tokens = set(texto_norm.split())

    for frase, peso in FRASES_AVISO_ALTA.items():
        frase_norm = normalizar_texto(frase)
        if frase_norm and frase_norm in texto_norm:
            score += peso
            motivos.append(f"frase correctiva: {frase}")

    for palabra, peso in PALABRAS_AVISO_ALTA.items():
        palabra_norm = normalizar_texto(palabra)
        if palabra_norm in tokens:
            score += peso
            motivos.append(f"palabra correctiva: {palabra}")

    for palabra, peso in PALABRAS_PREVENTIVO.items():
        palabra_norm = normalizar_texto(palabra)
        if palabra_norm and palabra_norm in texto_norm:
            score += peso
            motivos.append(f"patrón preventivo: {palabra}")

    return score, motivos


def score_numero_ot_vs_aviso(valor_ot):
    """
    Evalúa el número colocado en OT/Grafo usando patrón histórico.
    Score positivo = más sospecha de Aviso colocado como OT.

    Regla afinada:
    - 200, 300 y 400 se tratan como patrones probables de OT.
    - Solo los 100 bajos entran como zona realmente sospechosa de Aviso SAP.
    """
    n = limpiar_ot(valor_ot)

    if not n:
        return 0, "OT vacía."

    if not n.isdigit():
        return 80, "OT contiene caracteres no numéricos."

    if len(n) not in (8, 9):
        return 80, "OT no tiene 8 o 9 dígitos."

    # En el histórico, 300 y 400 son casi siempre OT.
    if n.startswith(("300", "400")):
        return -40, "Número con patrón muy probable de OT."

    # 200xxxxx es principalmente OT correctiva.
    # No debe observarse solo porque el texto diga falla/correctivo.
    if n.startswith("200"):
        return -35, "Número con patrón probable de OT correctiva."

    # Zona de mayor confusión: números 100 bajos.
    if n.startswith("100"):
        try:
            num = int(n)

            if len(n) == 8 and AVISO_RANGO_100_MIN <= num <= AVISO_RANGO_100_MAX:
                return 35, "Número dentro del rango típico de Avisos SAP."

            if num > AVISO_RANGO_100_MAX:
                return -20, "Número 100 fuera del rango típico de Avisos SAP."
        except Exception:
            pass

    if not n.startswith(("100", "200", "300", "400")):
        return 70, "Número no inicia con patrón típico de OT."

    return 0, "Patrón numérico neutro."


def evaluar_ot_vs_aviso(ot_grafo, aviso, actividad, tipo_mant="", condicion=""):
    """
    Detecta probable Aviso colocado como OT o Aviso informado sin OT.

    No usa tabla maestra SAP; usa score por:
    - patrón numérico,
    - texto correctivo/preventivo,
    - posición del dato: OT vacía / Aviso lleno.

    Devuelve dict con aplica/nivel/tipo/sugerencia.
    """
    ot = limpiar_ot(ot_grafo)
    av = limpiar_ot(aviso)
    texto = f"{actividad or ''} {tipo_mant or ''} {condicion or ''}"

    score_txt, motivos_txt = score_texto_aviso(texto)
    motivos = []

    # Caso crítico: el proveedor informó Aviso/COD PM, pero dejó OT vacía.
    if not ot and av:
        score = 45
        motivos.append("Aviso/COD PM informado con OT vacía")

        if av.isdigit() and len(av) in (8, 9) and av.startswith("100"):
            try:
                av_num = int(av)
                if len(av) == 8 and AVISO_RANGO_100_MIN <= av_num <= AVISO_RANGO_100_MAX:
                    score += 20
                    motivos.append("Aviso dentro del rango típico de Avisos SAP")
            except Exception:
                pass

        score += score_txt
        motivos.extend(motivos_txt[:4])
        score = max(0, min(score, 100))

        # Si el aviso está informado pero la actividad parece preventiva o
        # la sospecha es baja, no generamos observación. Evita ruido como:
        # inspecciones / preventivos con COD PM informado y OT aún vacía.
        if score < 45:
            return {
                "aplica": False,
                "nivel": "",
                "score": score,
                "campo": "",
                "valor": "",
                "tipo_observacion": "",
                "sugerencia": "",
                "motivos": motivos,
            }

        if score >= 70:
            return {
                "aplica": True,
                "nivel": "ADVERTENCIA",
                "score": score,
                "campo": "cod_pm_aviso",
                "valor": av,
                "tipo_observacion": "Probable aviso correctivo informado sin OT asociada.",
                "sugerencia": (
                    "Se informó un Aviso/COD PM, pero la OT está vacía. "
                    "La descripción tiene patrón correctivo. Verificar en SAP si el aviso ya tiene una OT asociada; "
                    "si existe, completar el campo OT/Grafo."
                ),
                "motivos": motivos,
            }

        return {
            "aplica": True,
            "nivel": "ADVERTENCIA",
            "score": score,
            "campo": "cod_pm_aviso",
            "valor": av,
            "tipo_observacion": "Aviso/COD PM informado sin OT asociada.",
            "sugerencia": (
                "Se informó Aviso/COD PM, pero el campo OT/Grafo está vacío. "
                "Verificar en SAP si el aviso ya tiene una OT asociada."
            ),
            "motivos": motivos,
        }

    # Si OT y Aviso son iguales, es sospechoso aunque el número tenga formato correcto.
    if ot and av and ot == av:
        return {
            "aplica": True,
            "nivel": "ADVERTENCIA",
            "score": 75,
            "campo": "ot_grafo",
            "valor": ot,
            "tipo_observacion": "Mismo número informado como OT y Aviso/COD PM.",
            "sugerencia": (
                "El mismo número aparece en OT/Grafo y en Aviso/COD PM. "
                "Verificar en SAP cuál corresponde realmente a la OT y cuál al aviso."
            ),
            "motivos": ["OT y Aviso/COD PM tienen el mismo valor"],
        }

    # Caso OT llena: evaluar si parece aviso.
    if ot:
        score_num, motivo_num = score_numero_ot_vs_aviso(ot)
        score = score_num + score_txt
        motivos.append(motivo_num)

        # Regla de exclusión:
        # Si el número informado como OT inicia con 200, 300 o 400,
        # no lo tratamos como posible Aviso SAP. En el histórico estos
        # patrones corresponden principalmente a OT, especialmente 200 como OT correctiva.
        # Esto evita falsos positivos como:
        # OT 20002260 + actividad "FALLA..." => debe ser probable OT, no aviso.
        if ot.startswith(("200", "300", "400")):
            return {
                "aplica": False,
                "nivel": "",
                "score": 0,
                "campo": "",
                "valor": "",
                "tipo_observacion": "",
                "sugerencia": "",
                "motivos": [],
            }

        motivos.extend(motivos_txt[:4])

        tipo_norm = normalizar_texto(tipo_mant)
        condicion_norm = normalizar_texto(condicion)
        texto_norm = normalizar_texto(texto)

        if any(p in tipo_norm for p in ["cond", "correct", "corr"]):
            score += 10
            motivos.append("Tipo de mantenimiento correctivo/condicional")

        if any(p in tipo_norm for p in ["prev", "mantto", "manto", "mantenimiento"]):
            score -= 12
            motivos.append("Tipo de mantenimiento con patrón preventivo")

        # Si dice correctivo o condición de falla, sube ligeramente.
        if any(p in texto_norm for p in ["correctivo", "condicional", "emergencia"]):
            score += 10
            motivos.append("Texto sugiere atención correctiva/condicional")

        score = max(0, min(score, 100))

        if score >= 70:
            return {
                "aplica": True,
                "nivel": "ADVERTENCIA",
                "score": score,
                "campo": "ot_grafo",
                "valor": ot,
                "tipo_observacion": "Alta probabilidad de Aviso colocado en campo OT.",
                "sugerencia": (
                    "El número tiene patrón similar a Aviso SAP y la descripción parece correctiva. "
                    "Verificar en SAP si corresponde a Aviso o a OT. Si es Aviso, reemplazar por la OT asociada."
                ),
                "motivos": motivos,
            }

        if score >= 40:
            return {
                "aplica": True,
                "nivel": "ADVERTENCIA",
                "score": score,
                "campo": "ot_grafo",
                "valor": ot,
                "tipo_observacion": "Número OT con patrón similar a Aviso SAP.",
                "sugerencia": (
                    "El número informado como OT cae en una zona típica de Avisos SAP. "
                    "Validar si corresponde realmente a OT/Grafo o si debe reemplazarse por la OT asociada."
                ),
                "motivos": motivos,
            }

    return {
        "aplica": False,
        "nivel": "",
        "score": 0,
        "campo": "",
        "valor": "",
        "tipo_observacion": "",
        "sugerencia": "",
        "motivos": [],
    }



# ============================================================
# AUTOCOMPLETADO INTELIGENTE DE CONDICIÓN
# ============================================================

def inferir_condicion_cond(row):
    """
    Infere E/S o F/S solo para actividades tipo COND con condición vacía.

    Regla conservadora:
    - F/S: intervención física, falla/fuga/inoperativo, cambio/reemplazo/montaje.
    - E/S: ajuste, revisión, evaluación, inspección, verificación, medición, calibración.
    - Si no hay señal clara, no infiere.
    """
    tipo_mant = normalizar_texto(row.get("tipo_mant", ""))
    condicion_actual = row.get("condicion", "")

    if "cond" not in tipo_mant:
        return ""

    if not esta_vacio(condicion_actual):
        return ""

    texto = normalizar_texto(
        " ".join([
            convertir_a_str_seguro(row.get("motivo", "")),
            convertir_a_str_seguro(row.get("sistema", "")),
            convertir_a_str_seguro(row.get("equipo", "")),
            convertir_a_str_seguro(row.get("observacion", "")),
            convertir_a_str_seguro(row.get("texto_explicativo", "")),
        ])
    )

    if not texto:
        return ""

    frases_fs = [
        "fuera de servicio",
        "fuga de aceite",
        "fuga aceite",
        "fuga de agua",
        "fuga agua",
        "fuga de aire",
        "fuga aire",
        "bajo nivel",
        "baja presion",
        "baja presión",
        "no comunica",
        "no responde",
        "sin comunicacion",
        "sin comunicación",
        "sin control",
        "no opera",
        "no operativo",
        "control deshabilitado",
        "comando deshabilitado",
        "equipo deshabilitado",
        "bloqueado por falla",
        "bloqueada por falla",
        "cambio de",
        "reemplazo de",
        "desmontaje de",
        "montaje de",
        "intervencion de",
        "intervención de",
    ]

    palabras_fs = {
        "falla",
        "fallo",
        "inoperativo",
        "fuga",
        "filtracion",
        "filtración",
        "rotura",
        "perdida",
        "pérdida",
        "trip",
        "alarma",
        "desgaste",
        "cambio",
        "reemplazo",
        "montaje",
        "desmontaje",
        "intervencion",
        "intervención",
        "averia",
        "avería",
        "deshabilitado",
        "deshabilitada",
        "bloqueado",
        "bloqueada",
        "desactivado",
        "desactivada",
        "dañado",
        "danado",
        "corregir",
        "correccion",
        "corrección",
    }

    frases_es = [
        "ajuste de",
        "ajuste presion",
        "ajuste presión",
        "revision de",
        "revisión de",
        "evaluacion de",
        "evaluación de",
        "inspeccion de",
        "inspección de",
        "verificacion de",
        "verificación de",
        "medicion de",
        "medición de",
        "calibracion de",
        "calibración de",
        "prueba de",
        "pruebas de",
        "monitoreo de",
    ]

    palabras_es = {
        "ajuste",
        "revision",
        "revisión",
        "evaluacion",
        "evaluación",
        "inspeccion",
        "inspección",
        "verificacion",
        "verificación",
        "medicion",
        "medición",
        "calibracion",
        "calibración",
        "prueba",
        "pruebas",
        "monitoreo",
        "diagnostico",
        "diagnóstico",
        "levantamiento",
    }

    tokens = set(texto.split())

    score_fs = 0
    score_es = 0

    for frase in frases_fs:
        if normalizar_texto(frase) in texto:
            score_fs += 3

    for palabra in palabras_fs:
        if normalizar_texto(palabra) in tokens:
            score_fs += 2

    for frase in frases_es:
        if normalizar_texto(frase) in texto:
            score_es += 3

    for palabra in palabras_es:
        if normalizar_texto(palabra) in tokens:
            score_es += 2

    # Prioridad a F/S cuando hay señal física/correctiva clara.
    if score_fs >= 2 and score_fs >= score_es:
        return "F/S"

    if score_es >= 2:
        return "E/S"

    return ""


def actualizar_datos_originales_condicion(datos_originales, condicion_inferida):
    """
    Actualiza datos_originales para que el generador del PMS único copie la
    condición inferida en la misma columna original, si esta fue detectada.
    """
    datos = json_seguro(datos_originales)

    if not isinstance(datos, dict):
        datos = {}

    datos["condicion_inferida_por_parser"] = True

    campos_detectados = datos.get("campos_detectados")
    if not isinstance(campos_detectados, dict):
        campos_detectados = {}
        datos["campos_detectados"] = campos_detectados

    condicion_original = campos_detectados.get("condicion")
    if condicion_original in [None, ""]:
        datos["condicion_original"] = condicion_original
    else:
        datos["condicion_original"] = condicion_original

    campos_detectados["condicion"] = condicion_inferida

    mapa_campos = datos.get("mapa_campos")
    columnas_excel = datos.get("columnas_excel")

    if isinstance(mapa_campos, dict) and isinstance(columnas_excel, dict):
        info_condicion = mapa_campos.get("condicion")

        if isinstance(info_condicion, dict):
            col_letra = info_condicion.get("col_letra")

            if col_letra:
                columnas_excel[col_letra] = condicion_inferida

    return datos


def autocompletar_condiciones_cond(df):
    """
    Completa la columna condicion para actividades tipo COND cuando está vacía
    y existe una inferencia clara. No genera observación; solo reduce ruido.
    """
    if df.empty or "condicion" not in df.columns or "tipo_mant" not in df.columns:
        return df

    df = df.copy()

    for idx, row in df.iterrows():
        condicion_inferida = inferir_condicion_cond(row)

        if condicion_inferida:
            df.at[idx, "condicion"] = condicion_inferida

            if "datos_originales" in df.columns:
                df.at[idx, "datos_originales"] = actualizar_datos_originales_condicion(
                    row.get("datos_originales", {}),
                    condicion_inferida,
                )

    return df


def validar_riesgo(valor):
    v = txt_upper(valor)

    if v in ["", "NAN", "NONE", "-", "NA", "N/A", "EMPTY"]:
        return "VACIO_VALIDO"

    if v == "X":
        return "VALIDO"

    if v in [
        "SI",
        "SÍ",
        "NO",
        "ALTO",
        "MEDIO",
        "BAJO",
        "CRITICO",
        "CRÍTICO",
        "CRITICAL",
    ]:
        return "VALOR_NO_ESTANDAR"

    return "INVALIDO"


def sistema_parece_unidad(valor):
    v = txt_upper(valor)

    if not v:
        return False

    if re.search(r"\bTG\s*-?\s*\d+\b", v):
        return True

    unidades = [
        "VENTANILLA",
        "SANTA ROSA",
        "MALACAS",
        "TG4",
        "TG5",
        "TG6",
        "TG7",
        "TG8",
        "DISTRIBUCION",
        "DISTRIBUCIÓN",
    ]

    return v in unidades


def validar_condicion(valor):
    v = txt_upper(valor)

    if v == "":
        return "VACIA"

    validos = [
        "E/S",
        "F/S",
        "E",
        "S",
        "OPERATIVO",
        "FUERA DE SERVICIO",
        "EN SERVICIO",
        "CON UNIDAD EN SERVICIO",
        "CON UNIDAD FUERA DE SERVICIO",
    ]

    if v in validos:
        return "VALIDA"

    return "NO_ESTANDAR"


def validar_recursos(valor):
    v = txt_upper(valor)

    if v in ["", "NA", "N/A", "NONE", "NAN", "-", "EMPTY"]:
        return "VACIO"

    try:
        float(v)
        return "VALIDO"
    except Exception:
        pass

    if re.fullmatch(r"\d+\s*P\s*\d+(\.\d+)?\s*[DH]", v):
        return "VALIDO"

    return "INVALIDO"


def agregar_obs(lista, row, campo, nivel, observacion, valor=None, sugerencia=None):
    inspector = row.get("inspector", "")
    rt = row.get("rt_terceros", "")

    if not esta_vacio(inspector):
        inspector_responsable = inspector
    else:
        inspector_responsable = rt

    try:
        fila_excel = int(row.get("fila_excel", 0) or 0)
    except Exception:
        fila_excel = 0

    central_original = convertir_a_str_seguro(row.get("central", ""))
    central_norm = convertir_a_str_seguro(row.get("central_norm", ""))

    lista.append({
        "proveedor": convertir_a_str_seguro(row.get("proveedor", "")),
        "hoja": convertir_a_str_seguro(row.get("hoja", "")),
        "fila_excel": fila_excel,
        "campo": campo,
        "nivel": nivel,
        "valor_detectado": convertir_a_str_seguro(
            valor if valor is not None else row.get(campo, "")
        ),
        "central": central_norm or normalizar_central_operativa(central_original),
        "central_original": central_original,
        "unidad": convertir_a_str_seguro(row.get("unidad", "")),
        "sistema": convertir_a_str_seguro(row.get("sistema", "")),
        "equipo": convertir_a_str_seguro(row.get("equipo", "")),
        "actividad": convertir_a_str_seguro(row.get("motivo", "")),
        "inspector_responsable": convertir_a_str_seguro(inspector_responsable),
        "tipo_observacion": observacion,
        "sugerencia": sugerencia or "",
    })


def generar_observaciones_forma(df):
    observaciones = []

    for _, row in df.iterrows():
        tipo_mant = txt_upper(row.get("tipo_mant", ""))

        # ============================================================
        # OT
        # ============================================================
        estado_ot, _ = validar_ot(row.get("ot_grafo", ""))

        if estado_ot == "VACIA":
            # Si existe Aviso/COD PM, se agregará una observación más específica
            # con evaluar_ot_vs_aviso(). Evitamos duplicar "Actividad sin OT".
            aviso_informado = not esta_vacio(row.get("cod_pm_aviso", ""))

            if not aviso_informado:
                agregar_obs(
                    observaciones,
                    row,
                    campo="ot_grafo",
                    nivel="ADVERTENCIA",
                    observacion="Actividad sin OT.",
                    valor=row.get("ot_grafo", ""),
                    sugerencia=(
                        "Completar la OT si ya fue generada. "
                        "Debe tener 8 o 9 dígitos y cumplir el patrón "
                        "100xxxxx, 100xxxxxx, 200xxxxx, 200xxxxxx, 300xxxxx, 300xxxxxx, 400xxxxx o 400xxxxxx."
                    ),
                )

        elif estado_ot == "PEDIDO_EN_CAMPO_OT":
            agregar_obs(
                observaciones,
                row,
                campo="ot_grafo",
                nivel="ADVERTENCIA",
                observacion="Número de pedido informado en columna OT/Grafo.",
                valor=row.get("ot_grafo", ""),
                sugerencia=(
                    "El número informado tiene patrón de pedido SAP (3500xxxx o 4500xxxx). "
                    "Debe trasladarse a la columna N° Pedido y confirmar si existe una OT real asociada."
                ),
            )

        elif estado_ot == "LONGITUD_INVALIDA":
            agregar_obs(
                observaciones,
                row,
                campo="ot_grafo",
                nivel="ERROR",
                observacion="OT con cantidad incorrecta de dígitos.",
                valor=row.get("ot_grafo", ""),
                sugerencia="Verificar OT. Debe tener 8 o 9 dígitos, por ejemplo 10006885 o 100012267.",
            )

        elif estado_ot == "PRIMER_DIGITO_INVALIDO":
            agregar_obs(
                observaciones,
                row,
                campo="ot_grafo",
                nivel="ERROR",
                observacion="OT inicia con un dígito no permitido.",
                valor=row.get("ot_grafo", ""),
                sugerencia=(
                    "La OT debe iniciar con 1, 2, 3 o 4 y cumplir el patrón "
                    "100xxxxx, 100xxxxxx, 200xxxxx, 200xxxxxx, 300xxxxx, 300xxxxxx, 400xxxxx o 400xxxxxx."
                ),
            )

        elif estado_ot == "PATRON_INVALIDO":
            agregar_obs(
                observaciones,
                row,
                campo="ot_grafo",
                nivel="ERROR",
                observacion="OT no cumple el patrón esperado.",
                valor=row.get("ot_grafo", ""),
                sugerencia=(
                    "La OT debe cumplir el patrón 100xxxxx, 100xxxxxx, "
                    "200xxxxx, 200xxxxxx, 300xxxxx, 300xxxxxx, 400xxxxx o 400xxxxxx. "
                    "No debe iniciar con 350, 450, 210, etc."
                ),
            )

        elif estado_ot == "INVALIDA":
            agregar_obs(
                observaciones,
                row,
                campo="ot_grafo",
                nivel="ERROR",
                observacion="Formato de OT no reconocido.",
                valor=row.get("ot_grafo", ""),
                sugerencia=(
                    "Completar una OT válida de 8 o 9 dígitos con patrón "
                    "100xxxxx, 100xxxxxx, 200xxxxx, 200xxxxxx, 300xxxxx, 300xxxxxx, 400xxxxx o 400xxxxxx."
                ),
            )

        # ============================================================
        # OT VS AVISO SAP - VALIDACIÓN INTELIGENTE
        # ============================================================
        # Se ejecuta solo cuando la OT está vacía o tiene formato válido.
        # Si la OT ya tiene error de formato, no duplicamos observaciones.
        if estado_ot in ["VACIA", "VALIDA"]:
            obs_ot_aviso = evaluar_ot_vs_aviso(
                ot_grafo=row.get("ot_grafo", ""),
                aviso=row.get("cod_pm_aviso", ""),
                actividad=row.get("motivo", ""),
                tipo_mant=row.get("tipo_mant", ""),
                condicion=row.get("condicion", ""),
            )

            if obs_ot_aviso.get("aplica"):
                motivos = obs_ot_aviso.get("motivos") or []
                score = obs_ot_aviso.get("score", 0)

                detalle_score = f" Score de sospecha: {score}/100."
                if motivos:
                    detalle_score += " Motivos: " + "; ".join(motivos[:3]) + "."

                agregar_obs(
                    observaciones,
                    row,
                    campo=obs_ot_aviso.get("campo") or "ot_grafo",
                    nivel=obs_ot_aviso.get("nivel") or "ADVERTENCIA",
                    observacion=obs_ot_aviso.get("tipo_observacion") or "Posible confusión entre OT y Aviso SAP.",
                    valor=obs_ot_aviso.get("valor") or row.get("ot_grafo", "") or row.get("cod_pm_aviso", ""),
                    sugerencia=(obs_ot_aviso.get("sugerencia") or "Validar OT/Aviso en SAP.") + detalle_score,
                )

        # ============================================================
        # UNIDAD
        # ============================================================
        if "unidad" in df.columns and esta_vacio(row.get("unidad", "")):
            agregar_obs(
                observaciones,
                row,
                campo="unidad",
                nivel="ADVERTENCIA",
                observacion="El campo Unidad / Grupo está vacío.",
                valor=row.get("unidad", ""),
                sugerencia=(
                    "Completar la unidad si aplica. "
                    "Si la actividad es común o transversal, usar COMUNES PLANTA "
                    "o el valor equivalente."
                ),
            )

        # ============================================================
        # CAMPOS OBLIGATORIOS TÉCNICOS
        # ============================================================
        campos_obligatorios_error = {
            "sistema": "Sistema",
            "equipo": "Sub sistema/equipo",
            "inspector": "Inspector Orygen",
            "rt_terceros": "RT terceros",
        }

        for campo, nombre in campos_obligatorios_error.items():
            if campo in df.columns and esta_vacio(row.get(campo, "")):
                agregar_obs(
                    observaciones,
                    row,
                    campo=campo,
                    nivel="ERROR",
                    observacion=f"El campo {nombre} está vacío.",
                    valor=row.get(campo, ""),
                    sugerencia=f"Completar el campo {nombre}.",
                )

        # ============================================================
        # CONDICIÓN
        # ============================================================
        if "condicion" in df.columns:
            condicion_vacia = esta_vacio(row.get("condicion", ""))

            if condicion_vacia and tipo_mant == "COND":
                agregar_obs(
                    observaciones,
                    row,
                    campo="condicion",
                    nivel="ADVERTENCIA",
                    observacion="Actividad COND sin condición informada.",
                    valor=row.get("condicion", ""),
                    sugerencia=(
                        "Completar la condición para actividades tipo COND. "
                        "Usar valores estándar como E/S, F/S, E o S."
                    ),
                )

            elif not condicion_vacia:
                estado_condicion = validar_condicion(row.get("condicion", ""))

                if estado_condicion == "NO_ESTANDAR":
                    agregar_obs(
                        observaciones,
                        row,
                        campo="condicion",
                        nivel="ADVERTENCIA",
                        observacion="La condición informada no tiene un formato estándar.",
                        valor=row.get("condicion", ""),
                        sugerencia="Usar valores estándar como E/S, F/S, E o S, según corresponda.",
                    )

        # ============================================================
        # SISTEMA CON POSIBLE UNIDAD
        # ============================================================
        if "sistema" in df.columns and sistema_parece_unidad(row.get("sistema", "")):
            agregar_obs(
                observaciones,
                row,
                campo="sistema",
                nivel="ERROR",
                observacion="El campo Sistema parece contener una unidad o central, no un sistema.",
                valor=row.get("sistema", ""),
                sugerencia=(
                    "Colocar el sistema correcto: Transformador, Protecciones, "
                    "Gasoducto, Agua de refrigeración, SCI, Diésel, etc."
                ),
            )

        # ============================================================
        # RIESGO CRÍTICO
        # ============================================================
        if "riesgo" in df.columns:
            estado_riesgo = validar_riesgo(row.get("riesgo", ""))

            if estado_riesgo == "VALOR_NO_ESTANDAR":
                agregar_obs(
                    observaciones,
                    row,
                    campo="riesgo",
                    nivel="ADVERTENCIA",
                    observacion="El campo Riesgo crítico tiene un valor no estándar.",
                    valor=row.get("riesgo", ""),
                    sugerencia="Usar solo X para marcar riesgo crítico. Dejar vacío si no aplica.",
                )

            elif estado_riesgo == "INVALIDO":
                agregar_obs(
                    observaciones,
                    row,
                    campo="riesgo",
                    nivel="ERROR",
                    observacion="El campo Riesgo crítico contiene un valor no válido.",
                    valor=row.get("riesgo", ""),
                    sugerencia="Usar solo X para riesgo crítico o dejar vacío.",
                )

        # ============================================================
        # RECURSOS
        # ============================================================
        if "recursos" in df.columns:
            estado_recursos = validar_recursos(row.get("recursos", ""))

            if estado_recursos == "INVALIDO":
                agregar_obs(
                    observaciones,
                    row,
                    campo="recursos",
                    nivel="ADVERTENCIA",
                    observacion="El campo recursos no tiene un formato reconocido.",
                    valor=row.get("recursos", ""),
                    sugerencia="Usar número simple o formato tipo 4P30D, 3P8H, 2P1.5H.",
                )

    df_obs = pd.DataFrame(observaciones)

    if df_obs.empty:
        df_obs = pd.DataFrame(columns=[
            "proveedor",
            "hoja",
            "fila_excel",
            "campo",
            "nivel",
            "valor_detectado",
            "central",
            "central_original",
            "unidad",
            "sistema",
            "equipo",
            "actividad",
            "inspector_responsable",
            "tipo_observacion",
            "sugerencia",
        ])

    return df_obs


# ============================================================
# FUNCIÓN PRINCIPAL USADA POR main.py
# ============================================================

def preparar_datos_parser(
    ruta_excel,
    pms_archivo_id=None,
    archivo_info=None,
    central_presentada=None,
):
    """
    Función principal llamada desde main.py.

    Compatible con:
    preparar_datos_parser(ruta_excel)
    preparar_datos_parser(ruta_excel, central_presentada="C. T. Santa Rosa")
    preparar_datos_parser(ruta_excel, pms_archivo_id=..., archivo_info=...)
    """

    if central_presentada is None:
        central_presentada = obtener_central_presentada_desde_archivo_info(archivo_info)

    proveedor_archivo = ""
    semana_archivo = ""
    archivo_path = ""

    if isinstance(archivo_info, dict):
        proveedor_archivo = archivo_info.get("proveedor") or ""
        semana_archivo = archivo_info.get("semana") or ""
        archivo_path = archivo_info.get("archivo_path") or ""

    df_actividades, df_hojas = extraer_actividades(ruta_excel)

    if df_actividades.empty:
        detalle_observaciones = [{
            "nivel": "ERROR",
            "tipo_observacion": "No se detectaron actividades en el PMS.",
            "central": "",
            "unidad": "",
            "actividad": "",
            "inspector_responsable": "",
            "fila_excel": 0,
            "campo": "archivo",
            "valor_detectado": "",
            "sugerencia": "Verificar que el archivo tenga hojas visibles con encabezados reconocibles.",
        }]

        return {
            "pms_archivo_id": pms_archivo_id,
            "proveedor": proveedor_archivo,
            "semana": semana_archivo,
            "archivo_path": archivo_path,
            "actividades": 0,
            "actividades_detalle": [],
            "observaciones": 1,
            "detalle_observaciones": detalle_observaciones,
            "hojas": df_hojas.to_dict(orient="records"),
            "centrales_detectadas": [],
            "central_presentada": central_presentada or "",
            "central_presentada_norm": normalizar_central_operativa(central_presentada),
            "errores": 1,
            "advertencias": 0,
            "estado": "ERROR - SIN ACTIVIDADES DETECTADAS",
        }

    df_limpio = df_actividades.copy()

    columnas_minimas = [
        "proveedor",
        "hoja",
        "central_hoja_detectada",
        "central",
        "unidad",
        "sistema",
        "equipo",
        "motivo",
        "ot_grafo",
        "tipo_mant",
        "condicion",
        "riesgo",
        "inspector",
        "rt_terceros",
        "recursos",
        "hora_inicio",
        "hora_fin",
        "cod_pm_aviso",
        "pedido",
        "numero_pedido",
        "ot_original_informada",
        "datos_originales",
    ]

    for col in columnas_minimas:
        if col not in df_limpio.columns:
            if col == "datos_originales":
                df_limpio[col] = [{} for _ in range(len(df_limpio))]
            else:
                df_limpio[col] = ""

    def inferir_central_desde_hoja(row):
        central = row.get("central", "")
        hoja = row.get("hoja", "")
        central_hoja = row.get("central_hoja_detectada", "")

        # 1) Si la fila tiene columna Central, manda la fila.
        if not esta_vacio(central):
            return central

        # 2) Si el encabezado/contenido superior de la hoja declara central, usarlo.
        #    Esto es más confiable que el nombre de la pestaña.
        if not esta_vacio(central_hoja):
            return central_hoja

        # 3) Si el proveedor declaró central en la web y la fila no trae central,
        #    asumir esa central. Esto cubre casos donde la hoja está mal nombrada.
        if central_presentada:
            return central_presentada

        # 4) Fallback histórico: usar nombre de hoja solo si no hay otra señal.
        hoja_limpia = limpiar_valor(hoja)

        if "SANTA ROSA" in hoja_limpia:
            return "SANTA ROSA"

        if "VENTANILLA" in hoja_limpia:
            return "VENTANILLA"

        if "DISTRIBUCION" in hoja_limpia or "DISTRIBUCIÓN" in hoja_limpia:
            return "DISTRIBUCION"

        if hoja_limpia:
            return hoja_limpia

        return ""

    df_limpio["central"] = df_limpio.apply(inferir_central_desde_hoja, axis=1)

    df_limpio = df_limpio.dropna(
        subset=["central", "unidad", "sistema", "equipo", "motivo"],
        how="all",
    )

    for col in df_limpio.columns:
        if col not in ["fila_excel", "datos_originales"]:
            df_limpio[col] = df_limpio[col].apply(convertir_a_str_seguro)

    df_limpio["datos_originales"] = df_limpio["datos_originales"].apply(json_seguro)

    # Detectar pedidos SAP informados en columna Pedido o por error en OT/Grafo.
    # No se borra todavía la OT informada: se conserva para validación y trazabilidad.
    if "pedido" not in df_limpio.columns:
        df_limpio["pedido"] = ""

    df_limpio["ot_original_informada"] = df_limpio["ot_grafo"].apply(convertir_a_str_seguro)
    df_limpio["numero_pedido"] = df_limpio.apply(
        lambda row: (
            extraer_numero_pedido(row.get("pedido", ""))
            or extraer_numero_pedido(row.get("ot_grafo", ""))
        ),
        axis=1,
    )

    for idx, row in df_limpio.iterrows():
        pedido_detectado = extraer_numero_pedido(row.get("ot_grafo", ""))
        if pedido_detectado:
            df_limpio.at[idx, "datos_originales"] = actualizar_datos_originales_pedido_desde_ot(
                row.get("datos_originales", {}),
                pedido_detectado,
            )

    cols_dedup = [
        "central",
        "unidad",
        "sistema",
        "equipo",
        "motivo",
        "condicion",
        "riesgo",
        "area_solicitante",
        "inspector",
        "rt_terceros",
        "hora_inicio",
        "hora_fin",
    ]

    cols_dedup = [c for c in cols_dedup if c in df_limpio.columns]

    if cols_dedup:
        df_limpio = df_limpio.drop_duplicates(subset=cols_dedup)

    df_base = df_limpio.copy()

    df_base["unidad_norm"] = df_base["unidad"].apply(normalizar_unidad)
    df_base["sistema_norm"] = df_base["sistema"].apply(normalizar_sistema)
    df_base["equipo_norm"] = df_base["equipo"].apply(limpiar_valor)
    df_base["riesgo_norm"] = df_base["riesgo"].apply(limpiar_valor)
    df_base["central_norm"] = df_base["central"].apply(normalizar_central_operativa)
    df_base["activo_padre"] = df_base.apply(inferir_activo_padre, axis=1)

    df_base_validable = df_base[
        df_base.apply(es_fila_actividad_real, axis=1)
    ].copy()

    # Autocompleta condición E/S o F/S para actividades COND cuando el texto permite inferirlo.
    # Esto reduce observaciones menores y permite que el PMS único salga más completo.
    df_base_validable = autocompletar_condiciones_cond(df_base_validable)

    # Filtrar por la central declarada en la web.
    # Esto evita que se validen hojas de Ventanilla/Distribución cuando el proveedor
    # registró el PMS para Santa Rosa, y viceversa.
    central_declarada_norm_pre = normalizar_central_operativa(central_presentada)
    if central_declarada_norm_pre and "central_norm" in df_base_validable.columns:
        df_base_validable = df_base_validable[
            df_base_validable["central_norm"].apply(normalizar_central_operativa) == central_declarada_norm_pre
        ].copy()

    if df_base_validable.empty:
        detalle_observaciones = [{
            "nivel": "ERROR",
            "tipo_observacion": "Se detectaron encabezados, pero no actividades reales validables.",
            "central": "",
            "unidad": "",
            "actividad": "",
            "inspector_responsable": "",
            "fila_excel": 0,
            "campo": "archivo",
            "valor_detectado": "",
            "sugerencia": "Verificar que el PMS tenga filas de actividades con central, unidad, sistema, equipo o motivo.",
        }]

        return {
            "pms_archivo_id": pms_archivo_id,
            "proveedor": proveedor_archivo,
            "semana": semana_archivo,
            "archivo_path": archivo_path,
            "actividades": 0,
            "actividades_detalle": [],
            "observaciones": 1,
            "detalle_observaciones": detalle_observaciones,
            "hojas": df_hojas.to_dict(orient="records"),
            "centrales_detectadas": [],
            "central_presentada": central_presentada or "",
            "central_presentada_norm": normalizar_central_operativa(central_presentada),
            "errores": 1,
            "advertencias": 0,
            "estado": "ERROR - SIN ACTIVIDADES VALIDABLES",
        }

    df_obs = generar_observaciones_forma(df_base_validable)

    central_declarada_norm = normalizar_central_operativa(central_presentada)

    centrales_detectadas = sorted([
        c for c in df_base_validable["central_norm"].dropna().unique().tolist()
        if c
    ])

    if central_declarada_norm:
        centrales_fuera = [
            c for c in centrales_detectadas
            if c and c != central_declarada_norm
        ]

        nuevas_obs_central = []

        for central_fuera in centrales_fuera:
            df_fuera = df_base_validable[
                df_base_validable["central_norm"] == central_fuera
            ]

            cantidad_fuera = len(df_fuera)

            nuevas_obs_central.append({
                "proveedor": "",
                "hoja": "",
                "fila_excel": 0,
                "campo": "central",
                "nivel": "ADVERTENCIA",
                "valor_detectado": central_fuera,
                "central": central_fuera,
                "central_original": central_fuera,
                "unidad": "",
                "sistema": "",
                "equipo": "",
                "actividad": f"Se detectaron {cantidad_fuera} actividades fuera de la central declarada.",
                "inspector_responsable": "",
                "tipo_observacion": "El archivo contiene actividades de una central distinta a la declarada.",
                "sugerencia": (
                    f"El proveedor declaró {central_presentada}, pero el archivo contiene "
                    f"{cantidad_fuera} actividades asociadas a {central_fuera}. "
                    "Verificar si corresponde separar el PMS por central o confirmar el alcance."
                ),
            })

        if nuevas_obs_central:
            df_obs = pd.concat(
                [df_obs, pd.DataFrame(nuevas_obs_central)],
                ignore_index=True,
            )

    errores = len(df_obs[df_obs["nivel"] == "ERROR"])
    advertencias = len(df_obs[df_obs["nivel"] == "ADVERTENCIA"])

    if errores == 0 and advertencias == 0:
        estado = "CONFORME"
    elif errores == 0 and advertencias > 0:
        estado = "CONFORME CON OBSERVACIONES"
    elif errores <= 5:
        estado = "OBSERVADO - CORRECCIÓN MENOR"
    else:
        estado = "OBSERVADO - REQUIERE CORRECCIÓN"

    actividades_detalle = []

    for _, r in df_base_validable.iterrows():
        fila_excel = r.get("fila_excel", 0)

        try:
            fila_excel = int(fila_excel or 0)
        except Exception:
            fila_excel = 0

        datos_originales = json_seguro(r.get("datos_originales", {}))

        actividad_item = {
            "fila_excel": fila_excel,
            "central": convertir_a_str_seguro(r.get("central_norm", r.get("central", ""))),
            "unidad": convertir_a_str_seguro(r.get("unidad", "")),
            "sistema": convertir_a_str_seguro(r.get("sistema", "")),
            "equipo": convertir_a_str_seguro(r.get("equipo", "")),
            "actividad": convertir_a_str_seguro(r.get("motivo", "")),
            "ot_grafo": convertir_a_str_seguro(r.get("ot_grafo", "")),
            "numero_pedido": convertir_a_str_seguro(r.get("numero_pedido", "")),
            "ot_original_informada": convertir_a_str_seguro(r.get("ot_original_informada", "")),
            "tipo_mant": convertir_a_str_seguro(r.get("tipo_mant", "")),
            "condicion": convertir_a_str_seguro(r.get("condicion", "")),
            "riesgo": convertir_a_str_seguro(r.get("riesgo", "")),
            "inspector": convertir_a_str_seguro(r.get("inspector", "")),
            "rt_terceros": convertir_a_str_seguro(r.get("rt_terceros", "")),
            "datos_originales": datos_originales,
        }

        campos_extra = [
            "proveedor",
            "hoja",
            "cod_pm_aviso",
            "pedido",
            "numero_pedido",
            "ot_original_informada",
            "central_hoja_detectada",
            "area_solicitante",
            "recursos",
            "hora_inicio",
            "hora_fin",
            "dias",
            "fecha_inicio",
            "fecha_fin",
            "codigo_actividad",
            "texto_explicativo",
            "riesgo_existente",
            "riesgo_introducido",
            "observacion",
        ]

        for campo in campos_extra:
            if campo in r:
                actividad_item[campo] = convertir_a_str_seguro(r.get(campo, ""))

        actividades_detalle.append(actividad_item)

    detalle_observaciones = []

    for _, r in df_obs.iterrows():
        fila_excel = r.get("fila_excel", 0)

        try:
            fila_excel = int(fila_excel or 0)
        except Exception:
            fila_excel = 0

        detalle_observaciones.append({
            "nivel": convertir_a_str_seguro(r.get("nivel", "")),
            "tipo_observacion": convertir_a_str_seguro(r.get("tipo_observacion", "")),
            "central": convertir_a_str_seguro(r.get("central", "")),
            "unidad": convertir_a_str_seguro(r.get("unidad", "")),
            "actividad": convertir_a_str_seguro(r.get("actividad", "")),
            "inspector_responsable": convertir_a_str_seguro(r.get("inspector_responsable", "")),
            "fila_excel": fila_excel,
            "campo": convertir_a_str_seguro(r.get("campo", "")),
            "valor_detectado": convertir_a_str_seguro(r.get("valor_detectado", "")),
            "sugerencia": convertir_a_str_seguro(r.get("sugerencia", "")),
        })

    return {
        "pms_archivo_id": pms_archivo_id,
        "proveedor": proveedor_archivo,
        "semana": semana_archivo,
        "archivo_path": archivo_path,

        # Resumen numérico para pms_archivos
        "actividades": int(len(actividades_detalle)),
        "observaciones": int(len(detalle_observaciones)),
        "errores": int(errores),
        "advertencias": int(advertencias),
        "estado": estado,

        # Detalle para insertar en tablas
        "actividades_detalle": actividades_detalle,
        "detalle_actividades": actividades_detalle,
        "actividades_data": actividades_detalle,
        "detalle_observaciones": detalle_observaciones,
        "observaciones_detalle": detalle_observaciones,
        "observaciones_data": detalle_observaciones,

        # Metadata
        "hojas": df_hojas.to_dict(orient="records"),
        "centrales_detectadas": centrales_detectadas,
        "central_presentada": central_presentada or "",
        "central_presentada_norm": central_declarada_norm,
    }
