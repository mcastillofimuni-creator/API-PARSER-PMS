import os
import re
import json
import tempfile
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import column_index_from_string, get_column_letter


PLANTILLA_NOMBRE = "PROGRAMA SEMANAL PLANTILLA.xlsx"

CENTRALES_VALIDAS = {
    "C. T. SANTA ROSA": "SANTA ROSA",
    "CT SANTA ROSA": "SANTA ROSA",
    "C.T. SANTA ROSA": "SANTA ROSA",
    "C T SANTA ROSA": "SANTA ROSA",
    "SANTA ROSA": "SANTA ROSA",
    "STA ROSA": "SANTA ROSA",

    "C.C. VENTANILLA": "VENTANILLA",
    "CC VENTANILLA": "VENTANILLA",
    "C.C VENTANILLA": "VENTANILLA",
    "C C VENTANILLA": "VENTANILLA",
    "VENTANILLA": "VENTANILLA",
}

CENTRAL_LABEL = {
    "SANTA ROSA": "C. T. Santa Rosa",
    "VENTANILLA": "C.C. Ventanilla",
}

MESES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}

DIAS_CORTOS = ["SÁB", "DOM", "LUN", "MAR", "MIÉ", "JUE", "VIE"]

# Columnas de programación dentro de la plantilla.
# U/V son Inicio/Fin y W:AC son Sáb-Vie.
COL_INICIO_PROG = "U"
COL_FIN_PROG = "V"
COL_DIAS_SEMANA = ["W", "X", "Y", "Z", "AA", "AB", "AC"]

# Colores de resaltado usados por el consolidado.
# Amarillo: fila con Riesgo Crítico = X. Azul: celda de día programado.
FILL_RIESGO_CRITICO = PatternFill(fill_type="solid", fgColor="FFFF00")
FILL_DIA_PROGRAMADO = PatternFill(fill_type="solid", fgColor="44546A")
FONT_DIA_PROGRAMADO = Font(color="FFFFFF", bold=True)

# Rango de columnas que se copiará desde datos_originales.columnas_excel.
# Según tu plantilla, el cuerpo relevante empieza en C y llega hasta AU.
COL_INICIO_DATOS = "C"
COL_FIN_DATOS = "AU"

FILA_INICIO_DATOS = 10
FILA_MODELO_ESTILO = 10

# Columnas fallback si datos_originales no trae alguna columna.
COLS_FALLBACK = {
    "ot_grafo": "C",
    "central": "D",
    "unidad": "E",
    "sistema": "F",
    "equipo": "G",
    "cod_pm_aviso": "H",
    "pedido": "I",
    "tipo_mant": "K",
    "condicion": "L",
    "riesgo": "M",
    "area_solicitante": "N",
    "inspector": "O",
    "rt_terceros": "P",
    "recursos": "Q",
    "hora_inicio": "R",
    "hora_fin": "S",
    "empresa": "AD",
    "actividad": "AF",
}


# Orden prioritario solicitado para el consolidado.
# Las empresas no listadas quedan debajo automáticamente.
ORDEN_EMPRESAS = [
    "MAGNEX",
    "JMI",
    "ULLOA",
]

UNIDADES_SANTA_ROSA_VALIDAS = {"COMUNES", "TG5", "TG6", "TG7", "TG8"}



# Catálogos normalizados para el consolidado.
# Se usan como una capa de criterio humano antes de escribir el PMS único.
AREAS_VALIDAS = {"MANTENIMIENTO", "OPERACIONES", "HSE&Q"}
TIPOS_MANT_VALIDOS = {"PREV", "CORR", "PROY"}
CONDICIONES_VALIDAS = {"E/S", "F/S"}

PROVEEDORES_HSEQ = {
    "ULLOA",
    "SGS",
    "LENOR",
}

INSPECTOR_OPERACIONES = "Y.ESPINOZA"

INSPECTORES = {
    "SANTA ROSA": {
        "MECANICA": ["C.LUQUE", "J.CACERES"],
        "ELECTRICIDAD": ["M.CASTILLO", "R.SANDOVAL"],
        "IC": ["P.GARCIA", "C.HUAYNATE", "M.TASAYCO"],
    },
    "VENTANILLA": {
        "MECANICA": ["F.CASTRO", "D.HINOSTROZA", "F.ROJAS"],
        "ELECTRICIDAD": ["F.SARMIENTO", "E.SALINAS"],
        "IC": ["C.ESPINOZA", "M.GOMEZ", "V.ESPIRITU"],
    },
}

ALIAS_INSPECTORES = {
    # Santa Rosa - Mecánica
    "CARLOS LUQUE": "C.LUQUE",
    "C LUQUE": "C.LUQUE",
    "C.LUQUE": "C.LUQUE",
    "JULIO CACERES": "J.CACERES",
    "JULIO CÁCERES": "J.CACERES",
    "J CACERES": "J.CACERES",
    "J.CACERES": "J.CACERES",
    # Santa Rosa - Electricidad
    "MANUEL CASTILLO": "M.CASTILLO",
    "M CASTILLO": "M.CASTILLO",
    "M.CASTILLO": "M.CASTILLO",
    "RICHARD SANDOVAL": "R.SANDOVAL",
    "R SANDOVAL": "R.SANDOVAL",
    "R.SANDOVAL": "R.SANDOVAL",
    # Santa Rosa - I&C
    "PERCY GARCIA": "P.GARCIA",
    "PERCY GARCÍA": "P.GARCIA",
    "P GARCIA": "P.GARCIA",
    "P.GARCIA": "P.GARCIA",
    "CESAR HUAYNATE": "C.HUAYNATE",
    "CÉSAR HUAYNATE": "C.HUAYNATE",
    "C HUAYNATE": "C.HUAYNATE",
    "C.HUAYNATE": "C.HUAYNATE",
    "MIGUEL TASAYCO": "M.TASAYCO",
    "M TASAYCO": "M.TASAYCO",
    "M.TASAYCO": "M.TASAYCO",
    # Ventanilla - Mecánica
    "FIDEL CASTRO": "F.CASTRO",
    "F CASTRO": "F.CASTRO",
    "F.CASTRO": "F.CASTRO",
    "DANIEL HINOSTROZA": "D.HINOSTROZA",
    "D HINOSTROZA": "D.HINOSTROZA",
    "D.HINOSTROZA": "D.HINOSTROZA",
    "FARID ROJAS": "F.ROJAS",
    "F ROJAS": "F.ROJAS",
    "F.ROJAS": "F.ROJAS",
    # Ventanilla - Electricidad
    "FERNANDO SARMIENTO": "F.SARMIENTO",
    "F SARMIENTO": "F.SARMIENTO",
    "F.SARMIENTO": "F.SARMIENTO",
    "EDUARDO SALINAS": "E.SALINAS",
    "E SALINAS": "E.SALINAS",
    "E.SALINAS": "E.SALINAS",
    # Ventanilla - I&C
    "CESAR ESPINOZA": "C.ESPINOZA",
    "CÉSAR ESPINOZA": "C.ESPINOZA",
    "C ESPINOZA": "C.ESPINOZA",
    "C.ESPINOZA": "C.ESPINOZA",
    "MIGUEL GOMEZ": "M.GOMEZ",
    "MIGUEL GÓMEZ": "M.GOMEZ",
    "M GOMEZ": "M.GOMEZ",
    "M.GOMEZ": "M.GOMEZ",
    "VICTOR ESPIRITU": "V.ESPIRITU",
    "VÍCTOR ESPÍRITU": "V.ESPIRITU",
    "V ESPIRITU": "V.ESPIRITU",
    "V.ESPIRITU": "V.ESPIRITU",
    # Operaciones
    "YERICO ESPINOZA": INSPECTOR_OPERACIONES,
    "Y ESPINOZA": INSPECTOR_OPERACIONES,
    "Y.ESPINOZA": INSPECTOR_OPERACIONES,
}


def normalizar_texto(valor):
    if valor is None:
        return ""

    texto = str(valor).strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto



def quitar_acentos_basico(valor):
    texto = normalizar_texto(valor)
    reemplazos = {
        "Á": "A", "É": "E", "Í": "I", "Ó": "O", "Ú": "U", "Ü": "U", "Ñ": "N",
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ü": "u", "ñ": "n",
    }

    for a, b in reemplazos.items():
        texto = texto.replace(a, b)

    return texto


def normalizar_empresa_consolidado(valor):
    """
    Normaliza nombres equivalentes de proveedor para ordenar y consolidar.

    Regla solicitada:
    - JMI y JM INGENIEROS se consideran la misma empresa: JMI.
    """
    empresa = quitar_acentos_basico(valor).upper()
    empresa = re.sub(r"[^A-Z0-9 ]+", " ", empresa)
    empresa = re.sub(r"\s+", " ", empresa).strip()

    if not empresa:
        return ""

    if empresa == "JMI" or "JM INGENIEROS" in empresa or "JM INGENIERIA" in empresa:
        return "JMI"

    if "MAGNEX" in empresa:
        return "MAGNEX"

    if "ULLOA" in empresa:
        return "ULLOA"

    return empresa


def clave_orden_empresa(actividad):
    """
    Ordena primero las empresas prioritarias y deja cualquier otra debajo.
    """
    proveedor = normalizar_empresa_consolidado(actividad.get("proveedor", ""))

    prioridad = {
        empresa: idx
        for idx, empresa in enumerate(ORDEN_EMPRESAS, start=1)
    }.get(proveedor, 999)

    try:
        fila_excel = int(actividad.get("fila_excel") or 0)
    except Exception:
        fila_excel = 0

    actividad_txt = normalizar_texto(actividad.get("actividad", "")).upper()
    ot = normalizar_texto(actividad.get("ot_grafo", "")).upper()

    return (
        prioridad,
        proveedor,
        fila_excel,
        ot,
        actividad_txt,
    )


def normalizar_unidad_santa_rosa(unidad, actividad=None):
    """
    Normaliza el campo GRUPO/UNIDAD para C. T. Santa Rosa.

    Valores permitidos:
    COMUNES, TG5, TG6, TG7, TG8.

    Reglas:
    - COMUNES PLANTA / PLANTA / ERM-* / BAJA / ALTA / SALA / CUBICULO -> COMUNES.
    - UTI o unidad no estándar: si el texto menciona TG5/TG6/TG7/TG8, usa esa TG.
      Si no hay TG clara, usa COMUNES.
    """
    unidad_txt = quitar_acentos_basico(unidad).upper()
    unidad_txt = re.sub(r"\s+", " ", unidad_txt).strip()

    partes = [unidad_txt]

    if isinstance(actividad, dict):
        for key in ["sistema", "equipo", "actividad", "motivo", "texto_explicativo", "observacion"]:
            partes.append(quitar_acentos_basico(actividad.get(key, "")).upper())

    texto_total = " ".join([p for p in partes if p])
    texto_total = re.sub(r"\s+", " ", texto_total).strip()

    for tg in ["TG5", "TG6", "TG7", "TG8"]:
        if re.search(rf"\b{tg}\b", unidad_txt):
            return tg

    # Casos explícitos de comunes.
    if (
        not unidad_txt
        or "COMUNES" in unidad_txt
        or "COMUN" in unidad_txt
        or "PLANTA" in unidad_txt
        or unidad_txt.startswith("ERM")
        or unidad_txt in {"BAJA", "ALTA", "SALA", "CUBICULO", "CUBÍCULO", "UTI"}
    ):
        # UTI suele requerir mirar el texto completo antes de mandarlo a COMUNES.
        if unidad_txt == "UTI":
            for tg in ["TG5", "TG6", "TG7", "TG8"]:
                if re.search(rf"\b{tg}\b", texto_total):
                    return tg

        return "COMUNES"

    # Si el proveedor puso algo no estándar pero el texto técnico contiene una TG,
    # usamos esa TG. Esto ayuda con actividades tipo Black Start TG5/TG6.
    for tg in ["TG5", "TG6", "TG7", "TG8"]:
        if re.search(rf"\b{tg}\b", texto_total):
            return tg

    # Para Santa Rosa, si no es una unidad válida, se considera actividad común.
    if unidad_txt not in UNIDADES_SANTA_ROSA_VALIDAS:
        return "COMUNES"

    return unidad_txt


def normalizar_central(valor):
    texto = normalizar_texto(valor).upper()
    texto = texto.replace(".", "")
    texto = re.sub(r"\s+", " ", texto)

    if "SANTA ROSA" in texto:
        return "SANTA ROSA"

    if "VENTANILLA" in texto:
        return "VENTANILLA"

    return CENTRALES_VALIDAS.get(texto, texto)


def parse_semana_inicio(semana):
    """
    Espera semana tipo: 2026-06-13.
    Ese día representa sábado.
    """
    try:
        return datetime.strptime(str(semana), "%Y-%m-%d").date()
    except Exception:
        raise ValueError("La semana debe tener formato YYYY-MM-DD, por ejemplo 2026-06-13.")


def fecha_larga(fecha):
    return f"{fecha.day:02d} DE {MESES[fecha.month]} DEL {fecha.year}"


def fecha_corta(fecha):
    return f"{fecha.day:02d}/{fecha.month:02d}/{fecha.year}"


def convertir_valor_excel(valor):
    """
    Limpia valores antes de escribirlos en Excel.
    Mantiene números y fechas cuando sea posible.
    """
    if valor is None:
        return None

    if isinstance(valor, str):
        v = valor.strip()
        if v.upper() in ["", "NULL", "NONE", "NAN"]:
            return None
        return valor

    return valor


def copiar_estilo_celda(origen, destino):
    if origen.has_style:
        destino.font = copy(origen.font)
        destino.fill = copy(origen.fill)
        destino.border = copy(origen.border)
        destino.alignment = copy(origen.alignment)
        destino.number_format = origen.number_format
        destino.protection = copy(origen.protection)

    if origen.hyperlink:
        destino._hyperlink = copy(origen.hyperlink)

    if origen.comment:
        destino.comment = copy(origen.comment)


def copiar_estilo_fila(ws, fila_origen, fila_destino, max_col):
    ws.row_dimensions[fila_destino].height = ws.row_dimensions[fila_origen].height

    for col in range(1, max_col + 1):
        copiar_estilo_celda(ws.cell(fila_origen, col), ws.cell(fila_destino, col))


def limpiar_area_datos(ws, fila_inicio=FILA_INICIO_DATOS):
    """
    Limpia valores del cuerpo de datos sin destruir formato.
    """
    for row in ws.iter_rows(min_row=fila_inicio, max_row=ws.max_row):
        for cell in row:
            cell.value = None


def obtener_valor(row, *keys, default=""):
    for key in keys:
        if key in row and row.get(key) not in [None, "", [], {}]:
            return row.get(key)
    return default


def parse_jsonb(valor):
    """
    Supabase puede devolver jsonb como dict o string.
    """
    if valor is None:
        return {}

    if isinstance(valor, dict):
        return valor

    if isinstance(valor, str):
        try:
            return json.loads(valor)
        except Exception:
            return {}

    return {}


def obtener_columnas_originales(actividad):
    """
    Lee datos_originales.columnas_excel.
    Estructura esperada:
    {
      "hoja": "...",
      "fila_excel": 10,
      "columnas_excel": {
        "A": "...",
        "B": "...",
        "C": "...",
        ...
      },
      "campos_detectados": {...}
    }
    """
    datos = parse_jsonb(actividad.get("datos_originales"))

    columnas = datos.get("columnas_excel")
    if isinstance(columnas, dict):
        return columnas

    return {}


def obtener_campos_detectados(actividad):
    datos = parse_jsonb(actividad.get("datos_originales"))

    campos = datos.get("campos_detectados")
    if isinstance(campos, dict):
        return campos

    return {}


def valor_original_o_fallback(actividad, columnas_originales, letra_columna, *fallback_keys):
    """
    Primero intenta usar la columna original del Excel.
    Si no existe, usa los campos planos de pms_actividades.
    """
    if letra_columna in columnas_originales:
        valor = columnas_originales.get(letra_columna)
        if valor not in [None, "", "NULL", "None", "nan"]:
            return convertir_valor_excel(valor)

    for key in fallback_keys:
        valor = actividad.get(key)
        if valor not in [None, "", "NULL", "None", "nan", [], {}]:
            return convertir_valor_excel(valor)

    return None


def actualizar_encabezado_semana(ws, semana_inicio):
    """
    Actualiza de forma robusta textos de fechas del encabezado.
    No depende de una sola celda exacta.
    """
    semana_fin = semana_inicio + timedelta(days=6)

    titulo = f"PROGRAMA SEMANAL DEL {fecha_larga(semana_inicio)} AL {fecha_larga(semana_fin)}"
    semana_texto = f"SEMANA DEL {fecha_corta(semana_inicio)} AL {fecha_corta(semana_fin)}"

    for row in ws.iter_rows(min_row=1, max_row=9):
        for cell in row:
            if isinstance(cell.value, str):
                texto = cell.value.upper()

                if "PROGRAMA SEMANAL" in texto and ("DEL" in texto or "SEMANA" in texto):
                    cell.value = titulo

                if "SEMANA" in texto and ("AL" in texto or "DEL" in texto):
                    cell.value = semana_texto

    # Buscar encabezados de días en el rango W:AC, pero sin asumir demasiado.
    # Si la plantilla tiene esos campos, los actualizamos.
    day_cols = COL_DIAS_SEMANA

    for i, col in enumerate(day_cols):
        fecha = semana_inicio + timedelta(days=i)

        try:
            ws[f"{col}8"] = fecha
            ws[f"{col}8"].number_format = "dd/mm/yyyy"
        except Exception:
            pass

        try:
            ws[f"{col}9"] = f"{DIAS_CORTOS[i]}\n{fecha.day:02d}/{fecha.month:02d}"
        except Exception:
            pass


def consultar_archivos_semana(supabase, semana, central_norm):
    """
    Trae archivos de la semana, filtra por central y se queda solo con
    el último archivo válido por proveedor.

    Esto evita que el consolidado jale versiones antiguas, pruebas o duplicados.
    """
    resp = (
        supabase.table("pms_archivos")
        .select(
            "id,semana,proveedor,expositor,central_presentada,central_presentada_norm,"
            "archivo_nombre,estado_validacion,errores,advertencias,actividades,"
            "observaciones,dias,fecha_carga"
        )
        .eq("semana", semana)
        .order("fecha_carga", desc=True)
        .execute()
    )

    archivos = resp.data or []

    archivos_validos = []

    for a in archivos:
        estado = normalizar_texto(a.get("estado_validacion", "")).upper()

        # Excluir cargas que nunca llegaron a validarse o siguen en proceso.
        if "ERROR API" in estado:
            continue

        if "VALIDANDO" in estado:
            continue

        if not a.get("id"):
            continue

        central_archivo = normalizar_central(
            a.get("central_presentada_norm")
            or a.get("central_presentada")
            or ""
        )

        if central_archivo != central_norm:
            continue

        proveedor = normalizar_empresa_consolidado(a.get("proveedor", ""))

        if not proveedor:
            proveedor = f"SIN_PROVEEDOR_{a.get('id')}"

        a["_proveedor_norm"] = proveedor
        archivos_validos.append(a)

    # Como vienen ordenados por fecha_carga desc, el primero de cada proveedor es el vigente.
    vigente_por_proveedor = {}

    for a in archivos_validos:
        proveedor = a["_proveedor_norm"]

        if proveedor not in vigente_por_proveedor:
            vigente_por_proveedor[proveedor] = a

    return vigente_por_proveedor

def consultar_actividades(supabase, semana, central_norm, archivos_vigentes):
    """
    Consulta actividades solo de los archivos vigentes por proveedor.
    Luego filtra filas pobres y elimina duplicados técnicos.
    """
    ids_vigentes = [a["id"] for a in archivos_vigentes.values() if a.get("id")]

    if not ids_vigentes:
        return []

    resp = (
        supabase.table("pms_actividades")
        .select("*")
        .eq("semana", semana)
        .eq("central", central_norm)
        .in_("pms_archivo_id", ids_vigentes)
        .order("proveedor")
        .order("fila_excel")
        .limit(5000)
        .execute()
    )

    actividades = resp.data or []

    actividades = [
        a for a in actividades
        if es_actividad_consolidable(a)
    ]

    actividades = quitar_duplicados_tecnicos(actividades)

    # Orden solicitado:
    # MAGNEX, JMI, ULLOA y luego cualquier otra empresa debajo.
    actividades = sorted(actividades, key=clave_orden_empresa)

    return actividades

def valor_limpio(valor):
    if valor is None:
        return ""

    texto = str(valor).strip()

    if texto.upper() in ["", "NULL", "NONE", "NAN", "EMPTY", "-", "--"]:
        return ""

    return texto


def obtener_columna_original(actividad, columna):
    """
    Devuelve valor desde datos_originales.columnas_excel.
    """
    columnas = obtener_columnas_originales(actividad)

    valor = columnas.get(columna)

    return valor_limpio(valor)


def es_actividad_consolidable(actividad):
    """
    Evita que entren filas pobres, subtítulos, residuos de formato
    o filas que solo tienen central/proveedor/actividad.

    Una actividad debe tener suficiente información técnica.
    """
    ot = valor_limpio(actividad.get("ot_grafo")) or obtener_columna_original(actividad, "C")
    central = valor_limpio(actividad.get("central")) or obtener_columna_original(actividad, "D")
    unidad = valor_limpio(actividad.get("unidad")) or obtener_columna_original(actividad, "E")
    sistema = valor_limpio(actividad.get("sistema")) or obtener_columna_original(actividad, "F")
    equipo = valor_limpio(actividad.get("equipo")) or obtener_columna_original(actividad, "G")
    tipo_mant = valor_limpio(actividad.get("tipo_mant")) or obtener_columna_original(actividad, "K")
    condicion = valor_limpio(actividad.get("condicion")) or obtener_columna_original(actividad, "L")
    inspector = valor_limpio(actividad.get("inspector")) or obtener_columna_original(actividad, "O")
    rt = valor_limpio(actividad.get("rt_terceros")) or obtener_columna_original(actividad, "P")
    actividad_txt = valor_limpio(actividad.get("actividad")) or obtener_columna_original(actividad, "AF")

    texto_total = " ".join([
        ot,
        central,
        unidad,
        sistema,
        equipo,
        tipo_mant,
        condicion,
        inspector,
        rt,
        actividad_txt,
    ]).upper()

    palabras_basura = [
        "TOTAL",
        "SUBTOTAL",
        "LEYENDA",
        "OBSERVACION",
        "OBSERVACIONES",
        "PROGRAMA SEMANAL",
        "SEMANA",
        "ELABORADO",
        "REVISADO",
        "APROBADO",
        "FIRMA",
    ]

    if any(p in texto_total for p in palabras_basura):
        return False

    campos_fuertes = [
        ot,
        unidad,
        sistema,
        equipo,
        tipo_mant,
        inspector,
        rt,
        actividad_txt,
    ]

    llenos = [c for c in campos_fuertes if valor_limpio(c)]

    # Regla práctica:
    # si no tiene al menos 4 campos técnicos fuertes, no debería ir al consolidado.
    if len(llenos) < 4:
        return False

    # Si solo tiene actividad pero no tiene sistema/equipo/unidad, probablemente es residuo.
    if actividad_txt and not unidad and not sistema and not equipo:
        return False

    return True


def clave_dedupe_actividad(actividad):
    """
    Clave técnica para evitar duplicados.
    Si una misma actividad viene repetida 4 veces por la misma carga,
    se queda solo una.
    """
    proveedor = normalizar_empresa_consolidado(actividad.get("proveedor", ""))
    central = valor_limpio(actividad.get("central")).upper()

    ot = valor_limpio(actividad.get("ot_grafo")) or obtener_columna_original(actividad, "C")
    unidad = valor_limpio(actividad.get("unidad")) or obtener_columna_original(actividad, "E")
    sistema = valor_limpio(actividad.get("sistema")) or obtener_columna_original(actividad, "F")
    equipo = valor_limpio(actividad.get("equipo")) or obtener_columna_original(actividad, "G")
    tipo_mant = valor_limpio(actividad.get("tipo_mant")) or obtener_columna_original(actividad, "K")
    actividad_txt = valor_limpio(actividad.get("actividad")) or obtener_columna_original(actividad, "AF")

    partes = [
        proveedor,
        central,
        ot,
        unidad,
        sistema,
        equipo,
        tipo_mant,
        actividad_txt,
    ]

    clave = " | ".join([normalizar_texto(p).upper() for p in partes])

    clave = re.sub(r"\s+", " ", clave).strip()

    return clave


def quitar_duplicados_tecnicos(actividades):
    """
    Elimina duplicados exactos o técnicamente equivalentes.
    """
    vistos = set()
    limpias = []

    for a in actividades:
        clave = clave_dedupe_actividad(a)

        if clave in vistos:
            continue

        vistos.add(clave)
        limpias.append(a)

    return limpias


def seleccionar_hoja(wb, central_norm):
    """
    Intenta seleccionar una hoja con nombre de la central.
    Si no existe, usa la hoja activa.
    """
    posibles = []

    if central_norm == "SANTA ROSA":
        posibles = ["SANTA ROSA", "STA ROSA", "CT SANTA ROSA", "C.T. SANTA ROSA"]

    elif central_norm == "VENTANILLA":
        posibles = ["VENTANILLA", "CC VENTANILLA", "C.C. VENTANILLA"]

    nombres = {ws.title.upper().strip(): ws.title for ws in wb.worksheets}

    for p in posibles:
        if p.upper() in nombres:
            return wb[nombres[p.upper()]]

    for ws in wb.worksheets:
        nombre = ws.title.upper()
        if central_norm in nombre:
            return ws

    return wb.active


def dejar_solo_hoja_objetivo(wb, ws_objetivo, central_norm):
    """
    Deja una sola hoja para evitar duplicidades.
    """
    ws_objetivo.title = central_norm

    for ws in list(wb.worksheets):
        if ws.title != ws_objetivo.title:
            wb.remove(ws)




def texto_base_actividad(actividad=None, *extras):
    """
    Junta la información disponible de la fila para inferir tipo, condición,
    área e inspector sin depender de una sola columna.
    """
    partes = []

    if isinstance(actividad, dict):
        for key in [
            "ot_grafo", "central", "unidad", "sistema", "equipo",
            "cod_pm_aviso", "pedido", "motivo", "actividad", "tipo_mant",
            "condicion", "riesgo", "area_solicitante", "inspector",
            "rt_terceros", "recursos", "texto_explicativo", "observacion",
        ]:
            partes.append(quitar_acentos_basico(actividad.get(key, "")).upper())

    for extra in extras:
        partes.append(quitar_acentos_basico(extra).upper())

    texto = " ".join([p for p in partes if p])
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def contiene_patron(texto, patrones):
    texto = quitar_acentos_basico(texto).upper()
    return any(re.search(p, texto) for p in patrones)


def normalizar_tipo_mant(valor, actividad=None, contexto_extra=""):
    """
    Normaliza TIPO MANT a PREV, CORR o PROY.
    Si el proveedor dejó el campo vacío o escribió un texto largo, se infiere
    por descripción de actividad/sistema/equipo.
    """
    actual = quitar_acentos_basico(valor).upper().strip()
    actual = re.sub(r"[^A-Z0-9/ ]+", " ", actual)
    actual = re.sub(r"\s+", " ", actual).strip()

    if actual in TIPOS_MANT_VALIDOS:
        return actual

    if actual in {"PREVENTIVO", "M1", "PM", "P M", "MANTO PREVENTIVO", "MANTTO PREVENTIVO"}:
        return "PREV"

    if actual in {"CORRECTIVO", "COND", "CONDICIONAL", "M2", "CM", "C M"}:
        return "CORR"

    if actual in {"PROYECTO", "PROYECTOS", "M3", "M4", "OBRA"}:
        return "PROY"

    texto = texto_base_actividad(actividad, contexto_extra, actual)

    patrones_proy = [
        r"\bPROY(ECTO)?\b", r"\bIMPLEMENTACION\b", r"\bIMPLEMENTAR\b",
        r"\bINSTALACION\b", r"\bINSTALAR\b", r"\bMONTAJE\b",
        r"\bNUEVO\b", r"\bNUEVA\b", r"\bOBRA\b", r"\bAMPLIACION\b",
        r"\bMODIFICACION\b", r"\bMEJORA\b", r"\bUPGRADE\b", r"\bCOMISIONAMIENTO\b",
    ]

    patrones_corr = [
        r"\bFALLA\b", r"\bFALLO\b", r"\bAVERIA\b", r"\bINOPERATIVO\b",
        r"\bINOPERATIVA\b", r"\bALARMA\b", r"\bFUGA\b", r"\bFILTRACION\b",
        r"\bPERDIDA\b", r"\bDESGASTE\b", r"\bROTURA\b", r"\bRUIDO\b",
        r"\bPASE\b", r"\bBAJA PRESION\b", r"\bBAJO NIVEL\b", r"\bTRIP\b",
        r"\bCORRECTIVO\b", r"\bREPARACION\b", r"\bREPARAR\b", r"\bREEMPLAZO\b",
        r"\bCAMBIO\b", r"\bCONTROL DESHABILITADO\b", r"\bNO OPERA\b",
    ]

    patrones_prev = [
        r"\bMANTTO\b", r"\bMANTO\b", r"\bMANTENIMIENTO\b", r"\bPREVENTIVO\b",
        r"\bINSPECCION\b", r"\bREVISION\b", r"\bVERIFICACION\b", r"\bPRUEBA\b",
        r"\bPRUEBAS\b", r"\bLIMPIEZA\b", r"\bCALIBRACION\b", r"\bMEDICION\b",
        r"\bLAVADO\b", r"\bANUAL\b", r"\bRUTINARIO\b", r"\bMONITOREO\b",
        r"\bTERMOGRAFIA\b", r"\bANALISIS\b",
    ]

    # Si hay palabras de proyecto explícitas, se considera proyecto salvo que el texto indique falla clara.
    if contiene_patron(texto, patrones_proy) and not contiene_patron(texto, patrones_corr):
        return "PROY"

    if contiene_patron(texto, patrones_corr):
        return "CORR"

    if contiene_patron(texto, patrones_proy):
        return "PROY"

    if contiene_patron(texto, patrones_prev):
        return "PREV"

    return actual if actual else "PREV"


def normalizar_condicion(valor, actividad=None, tipo_mant="", contexto_extra=""):
    """
    Normaliza CONDICIÓN a E/S o F/S.
    Respeta E/S y F/S si ya vienen correctamente.
    """
    actual = quitar_acentos_basico(valor).upper().strip()
    actual = actual.replace(" ", "")

    if actual in {"E/S", "ES", "E.S", "ENSERVICIO"}:
        return "E/S"

    if actual in {"F/S", "FS", "F.S", "FUERADESERVICIO"}:
        return "F/S"

    texto = texto_base_actividad(actividad, contexto_extra, tipo_mant, valor)

    patrones_fs = [
        r"\bFALLA\b", r"\bFUGA\b", r"\bFILTRACION\b", r"\bINOPERATIVO\b",
        r"\bINOPERATIVA\b", r"\bDESHABILITADO\b", r"\bDESHABILITADA\b",
        r"\bBLOQUEADO\b", r"\bBLOQUEADA\b", r"\bNO OPERA\b", r"\bCAMBIO\b",
        r"\bREEMPLAZO\b", r"\bDESMONTAJE\b", r"\bMONTAJE\b", r"\bINTERVENCION\b",
        r"\bLAVADO DE IGV\b", r"\bLAVADO MANUAL DE IGV\b", r"\bACOPLE\b",
        r"\bCORRECTIVO\b", r"\bCORR\b", r"\bPUESTA EN MARCHA\b",
    ]

    patrones_es = [
        r"\bAJUSTE\b", r"\bREVISION\b", r"\bEVALUACION\b", r"\bINSPECCION\b",
        r"\bVERIFICACION\b", r"\bMEDICION\b", r"\bMONITOREO\b", r"\bTERMOGRAFIA\b",
        r"\bANALISIS\b", r"\bPRUEBA FUNCIONAL\b", r"\bCALIBRACION\b",
    ]

    if contiene_patron(texto, patrones_fs):
        return "F/S"

    if contiene_patron(texto, patrones_es):
        return "E/S"

    tipo = quitar_acentos_basico(tipo_mant).upper()
    if tipo == "CORR":
        return "F/S"

    return "E/S"


def normalizar_area_solicitante(valor, actividad=None, proveedor=""):
    """
    Normaliza ÁREA SOLICITANTE a MANTENIMIENTO, OPERACIONES o HSE&Q.
    Usa proveedor y texto de actividad para detectar HSE&Q.
    """
    actual = quitar_acentos_basico(valor).upper().strip()
    actual = re.sub(r"[^A-Z0-9& ]+", " ", actual)
    actual = re.sub(r"\s+", " ", actual).strip()

    texto = texto_base_actividad(actividad, proveedor, actual)
    proveedor_norm = normalizar_empresa_consolidado(proveedor)

    if "OPERACION" in actual or "OPERACIONES" in actual or contiene_patron(texto, [r"\bOPERACIONES\b"]):
        return "OPERACIONES"

    patrones_hseq = [
        r"\bHSE\b", r"\bHSEQ\b", r"\bHSE&Q\b", r"\bSSOMA\b", r"\bSOMA\b",
        r"\bSEGURIDAD\b", r"\bSALUD\b", r"\bAMBIENTAL\b", r"\bAMBIENTE\b",
        r"\bMEDICO\b", r"\bMEDICA\b", r"\bERGONOMIA\b", r"\bERGONOMICO\b",
        r"\bRESIDUO\b", r"\bRESIDUOS\b", r"\bSOLIDOS\b", r"\bMONITOR(EO)? AMBIENTAL\b",
        r"\bMONITORES AMBIENTALES\b", r"\bRECOJO\b", r"\bDISPOSICION\b",
    ]

    if proveedor_norm in PROVEEDORES_HSEQ or contiene_patron(texto, patrones_hseq):
        return "HSE&Q"

    if "MANT" in actual or "MANTENIMIENTO" in actual:
        return "MANTENIMIENTO"

    return "MANTENIMIENTO"


def normalizar_inspector_existente(valor):
    inspector = quitar_acentos_basico(valor).upper().strip()
    inspector = re.sub(r"[^A-Z0-9./ ]+", " ", inspector)
    inspector = re.sub(r"\s+", " ", inspector).strip()

    if not inspector:
        return ""

    # Si vienen varios inspectores separados por '/', normaliza los que reconozca.
    partes = [p.strip() for p in re.split(r"/|,|;| Y ", inspector) if p.strip()]
    normalizados = []

    for parte in partes:
        parte_norm = parte.replace(" ", "") if "." in parte else parte
        candidato = ALIAS_INSPECTORES.get(parte) or ALIAS_INSPECTORES.get(parte_norm)
        if candidato and candidato not in normalizados:
            normalizados.append(candidato)

    if normalizados:
        return " / ".join(normalizados)

    return ""


def detectar_disciplina(actividad=None, contexto_extra=""):
    texto = texto_base_actividad(actividad, contexto_extra)

    patrones_ic = [
        r"\bINSTRUMENT", r"\bCONTROL\b", r"\bPLC\b", r"\bDCS\b", r"\bSCADA\b",
        r"\bRTU\b", r"\bTRANSMISOR\b", r"\bSENSOR\b", r"\bSWITCH\b", r"\bVALVULA DE CONTROL\b",
        r"\bVALVULA SOLENOIDE\b", r"\bPRESION\b", r"\bTEMPERATURA\b", r"\bCAUDAL\b",
        r"\bVIBRACION\b", r"\bPOSICION\b", r"\bACTUADOR\b", r"\bCONVERTIDOR\b",
    ]

    patrones_elec = [
        r"\bELECT", r"\bGENERADOR\b", r"\bTRANSFORMADOR\b", r"\bTRAFO\b",
        r"\bTABLERO\b", r"\bBREAKER\b", r"\bINTERRUPTOR\b", r"\bRELE\b", r"\bPROTECCION\b",
        r"\bUPS\b", r"\bBATERIA\b", r"\bCARGADOR\b", r"\bVARIADOR\b", r"\bEXCITACION\b",
        r"\bMOTOR ELECTRICO\b", r"\bMOTOR MT\b", r"\bMCC\b", r"\bCABLE\b", r"\bLUMINARIA\b",
        r"\bILUMINACION\b", r"\bPANEL SOLAR\b", r"\bPANELES SOLARES\b",
    ]

    patrones_mec = [
        r"\bMECAN", r"\bBOMBA\b", r"\bCOMPRESOR", r"\bVALVULA\b", r"\bFILTRO\b",
        r"\bTUBERIA\b", r"\bSKID\b", r"\bMOTOR DIESEL\b", r"\bIGV\b", r"\bLAVADO\b",
        r"\bGRUA\b", r"\bACOPLE\b", r"\bRODAMIENTO\b", r"\bREDUCTOR\b", r"\bLUBRICACION\b",
        r"\bACEITE\b", r"\bFUGA\b", r"\bENFRIAMIENTO\b", r"\bCALDERIN\b",
    ]

    # Prioridad: I&C para control/instrumentación; electricidad para equipos eléctricos;
    # mecánica al final para no capturar 'válvula de control' como mecánica.
    if contiene_patron(texto, patrones_ic):
        return "IC"
    if contiene_patron(texto, patrones_elec):
        return "ELECTRICIDAD"
    if contiene_patron(texto, patrones_mec):
        return "MECANICA"

    return "MECANICA"


def normalizar_inspector(valor, central_norm, area, actividad=None, contexto_extra=""):
    """
    Corrige el inspector usando biblioteca por central y disciplina.
    Si el área es OPERACIONES, siempre devuelve Y.ESPINOZA.
    """
    area_norm = quitar_acentos_basico(area).upper()

    if area_norm == "OPERACIONES":
        return INSPECTOR_OPERACIONES

    existente = normalizar_inspector_existente(valor)
    permitidos = []

    for lista in INSPECTORES.get(central_norm, {}).values():
        permitidos.extend(lista)

    if existente:
        # Si ya está dentro de la biblioteca de la central, respétalo.
        existentes = [x.strip() for x in existente.split("/")]
        if all(x in permitidos or x == INSPECTOR_OPERACIONES for x in existentes):
            return existente

    if area_norm == "HSE&Q":
        # No tenemos biblioteca HSE&Q de supervisores; si no hay válido, no inventamos.
        return existente or ""

    disciplina = detectar_disciplina(actividad, contexto_extra)
    opciones = INSPECTORES.get(central_norm, {}).get(disciplina, [])

    if opciones:
        return opciones[0]

    return existente or ""


def aplicar_normalizaciones_consolidadas(ws, fila, actividad, archivo_info=None):
    """
    Completa/normaliza Tipo Mant, Condición, Área Solicitante e Inspector.
    Esta función se ejecuta después de copiar la fila original y aplicar fallbacks,
    para que trabaje sobre los valores finales visibles en el Excel.
    """
    central_norm = normalizar_central(ws[f"{COLS_FALLBACK['central']}{fila}"].value)
    proveedor = ws[f"{COLS_FALLBACK['empresa']}{fila}"].value

    contexto_celdas = " ".join([
        normalizar_texto(ws[f"{COLS_FALLBACK['unidad']}{fila}"].value),
        normalizar_texto(ws[f"{COLS_FALLBACK['sistema']}{fila}"].value),
        normalizar_texto(ws[f"{COLS_FALLBACK['equipo']}{fila}"].value),
        normalizar_texto(ws[f"{COLS_FALLBACK['actividad']}{fila}"].value),
        normalizar_texto(ws[f"{COLS_FALLBACK['rt_terceros']}{fila}"].value),
        normalizar_texto(ws[f"{COLS_FALLBACK['recursos']}{fila}"].value),
        normalizar_texto(proveedor),
    ])

    # Área solicitante.
    area_actual = ws[f"{COLS_FALLBACK['area_solicitante']}{fila}"].value
    area_norm = normalizar_area_solicitante(area_actual, actividad, proveedor)
    ws[f"{COLS_FALLBACK['area_solicitante']}{fila}"] = area_norm

    # Tipo de mantenimiento.
    tipo_actual = ws[f"{COLS_FALLBACK['tipo_mant']}{fila}"].value
    tipo_norm = normalizar_tipo_mant(tipo_actual, actividad, contexto_celdas)
    ws[f"{COLS_FALLBACK['tipo_mant']}{fila}"] = tipo_norm

    # Condición.
    condicion_actual = ws[f"{COLS_FALLBACK['condicion']}{fila}"].value
    condicion_norm = normalizar_condicion(condicion_actual, actividad, tipo_norm, contexto_celdas)
    ws[f"{COLS_FALLBACK['condicion']}{fila}"] = condicion_norm

    # Inspector.
    inspector_actual = ws[f"{COLS_FALLBACK['inspector']}{fila}"].value
    inspector_norm = normalizar_inspector(inspector_actual, central_norm, area_norm, actividad, contexto_celdas)
    if inspector_norm:
        ws[f"{COLS_FALLBACK['inspector']}{fila}"] = inspector_norm



def parsear_fecha_para_excel(valor):
    """
    Convierte fechas tipo '2026-06-22 00:00:00' o '2026-06-22'
    a datetime/date para que Excel pueda mostrarlas como DD/MM/AAAA.
    """
    if valor is None:
        return None

    if isinstance(valor, datetime):
        return valor.date()

    texto = normalizar_texto(valor)

    if not texto:
        return valor

    formatos = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d",
        "%d/%m/%Y",
    ]

    for fmt in formatos:
        try:
            return datetime.strptime(texto, fmt).date()
        except Exception:
            pass

    return valor


def valor_marca_dia(valor):
    """
    Determina si una celda de día debe resaltarse.
    Ejemplos válidos: 1, 2, 3, '2', 'x'.
    """
    if valor is None:
        return False

    texto = normalizar_texto(valor)
    if texto.upper() in ["", "NULL", "NONE", "NAN", "EMPTY", "-", "--", "0"]:
        return False

    return True


def aplicar_formato_final_fila(ws, fila):
    """
    Ajustes finales después de copiar la fila original:
    - Inicio y Fin en formato DD/MM/AAAA.
    - Celdas de días programados con relleno azul.
    - Si Riesgo Crítico = X, resalta toda la fila en amarillo.
    """
    col_inicio = column_index_from_string(COL_INICIO_DATOS)
    col_fin = column_index_from_string(COL_FIN_DATOS)

    # 1) Fechas Inicio/Fin.
    for col in [COL_INICIO_PROG, COL_FIN_PROG]:
        celda = ws[f"{col}{fila}"]
        celda.value = parsear_fecha_para_excel(celda.value)
        if celda.value not in [None, ""]:
            celda.number_format = "dd/mm/yyyy"

    # 2) Riesgo crítico: si M tiene X, pintar toda la fila en amarillo.
    riesgo = normalizar_texto(ws[f"{COLS_FALLBACK['riesgo']}{fila}"].value).upper()
    riesgo_critico = bool(re.search(r"\bX\b", riesgo))

    if riesgo_critico:
        for col_idx in range(col_inicio, col_fin + 1):
            ws.cell(row=fila, column=col_idx).fill = copy(FILL_RIESGO_CRITICO)

    # 3) Marcas de programación diaria: mantenerlas azules, incluso si la fila es amarilla.
    for col in COL_DIAS_SEMANA:
        celda = ws[f"{col}{fila}"]
        if valor_marca_dia(celda.value):
            celda.fill = copy(FILL_DIA_PROGRAMADO)
            celda.font = copy(FONT_DIA_PROGRAMADO)

def escribir_fila_desde_original(ws, fila_destino, actividad, archivo_info):
    """
    Escribe una fila consolidada copiando el rango C:AU desde datos_originales.columnas_excel.

    Si alguna columna no viene en el JSON, usa fallback desde pms_actividades.
    """
    columnas_originales = obtener_columnas_originales(actividad)

    col_inicio = column_index_from_string(COL_INICIO_DATOS)
    col_fin = column_index_from_string(COL_FIN_DATOS)

    for col_idx in range(col_inicio, col_fin + 1):
        letra = get_column_letter(col_idx)
        valor = columnas_originales.get(letra)
        ws.cell(row=fila_destino, column=col_idx).value = convertir_valor_excel(valor)

    # Fallbacks y normalizaciones clave.
    proveedor = obtener_valor(
        actividad,
        "proveedor",
        default=archivo_info.get("proveedor", "") if archivo_info else "",
    )

    # Normalizar empresa para el consolidado.
    # JMI y JM INGENIEROS se muestran como JMI.
    col_empresa = COLS_FALLBACK["empresa"]
    proveedor_consolidado = normalizar_empresa_consolidado(proveedor)
    ws[f"{col_empresa}{fila_destino}"] = proveedor_consolidado or proveedor

    # Forzar central normalizada en D para que el consolidado quede limpio.
    ws[f"{COLS_FALLBACK['central']}{fila_destino}"] = obtener_valor(
        actividad,
        "central",
        default=normalizar_central(archivo_info.get("central_presentada", "")) if archivo_info else "",
    )

    # Si alguna columna principal vino vacía, rellenar con campos planos.
    ws[f"{COLS_FALLBACK['ot_grafo']}{fila_destino}"] = valor_original_o_fallback(
        actividad,
        columnas_originales,
        COLS_FALLBACK["ot_grafo"],
        "ot_grafo",
    )

    ws[f"{COLS_FALLBACK['unidad']}{fila_destino}"] = valor_original_o_fallback(
        actividad,
        columnas_originales,
        COLS_FALLBACK["unidad"],
        "unidad",
    )

    # Normalizar GRUPO/UNIDAD para C. T. Santa Rosa.
    central_actual = normalizar_central(ws[f"{COLS_FALLBACK['central']}{fila_destino}"].value)
    if central_actual == "SANTA ROSA":
        unidad_actual = ws[f"{COLS_FALLBACK['unidad']}{fila_destino}"].value
        ws[f"{COLS_FALLBACK['unidad']}{fila_destino}"] = normalizar_unidad_santa_rosa(
            unidad_actual,
            actividad,
        )

    ws[f"{COLS_FALLBACK['sistema']}{fila_destino}"] = valor_original_o_fallback(
        actividad,
        columnas_originales,
        COLS_FALLBACK["sistema"],
        "sistema",
    )

    ws[f"{COLS_FALLBACK['equipo']}{fila_destino}"] = valor_original_o_fallback(
        actividad,
        columnas_originales,
        COLS_FALLBACK["equipo"],
        "equipo",
    )

    ws[f"{COLS_FALLBACK['tipo_mant']}{fila_destino}"] = valor_original_o_fallback(
        actividad,
        columnas_originales,
        COLS_FALLBACK["tipo_mant"],
        "tipo_mant",
    )

    ws[f"{COLS_FALLBACK['condicion']}{fila_destino}"] = valor_original_o_fallback(
        actividad,
        columnas_originales,
        COLS_FALLBACK["condicion"],
        "condicion",
    )

    ws[f"{COLS_FALLBACK['area_solicitante']}{fila_destino}"] = valor_original_o_fallback(
        actividad,
        columnas_originales,
        COLS_FALLBACK["area_solicitante"],
        "area_solicitante",
    )

    ws[f"{COLS_FALLBACK['riesgo']}{fila_destino}"] = valor_original_o_fallback(
        actividad,
        columnas_originales,
        COLS_FALLBACK["riesgo"],
        "riesgo",
    )

    ws[f"{COLS_FALLBACK['inspector']}{fila_destino}"] = valor_original_o_fallback(
        actividad,
        columnas_originales,
        COLS_FALLBACK["inspector"],
        "inspector",
        "inspector_responsable",
    )

    ws[f"{COLS_FALLBACK['rt_terceros']}{fila_destino}"] = valor_original_o_fallback(
        actividad,
        columnas_originales,
        COLS_FALLBACK["rt_terceros"],
        "rt_terceros",
    )

    ws[f"{COLS_FALLBACK['actividad']}{fila_destino}"] = valor_original_o_fallback(
        actividad,
        columnas_originales,
        COLS_FALLBACK["actividad"],
        "actividad",
        "motivo",
    )

    aplicar_normalizaciones_consolidadas(ws, fila_destino, actividad, archivo_info)
    aplicar_formato_final_fila(ws, fila_destino)


def preparar_filas_destino(ws, total_actividades):
    """
    Prepara suficientes filas copiando el estilo de la fila modelo.
    No borra encabezados.
    """
    max_col = max(ws.max_column, column_index_from_string(COL_FIN_DATOS))

    if total_actividades <= 0:
        total_actividades = 1

    # Limpiar valores existentes.
    limpiar_area_datos(ws, fila_inicio=FILA_INICIO_DATOS)

    # Insertar filas si la plantilla no tiene suficientes.
    filas_actuales_desde_inicio = max(ws.max_row - FILA_INICIO_DATOS + 1, 1)

    if total_actividades > filas_actuales_desde_inicio:
        filas_extra = total_actividades - filas_actuales_desde_inicio
        ws.insert_rows(ws.max_row + 1, amount=filas_extra)

    # Copiar estilo de fila modelo a todas las filas destino.
    for idx in range(total_actividades):
        fila = FILA_INICIO_DATOS + idx
        copiar_estilo_fila(ws, FILA_MODELO_ESTILO, fila, max_col)


def obtener_numero_pms(semana):
    """
    Calcula el número PMS usando la semana ISO.
    Ejemplo:
    2026-06-20 corresponde a PMS 25.
    """
    semana_inicio = parse_semana_inicio(semana)
    return semana_inicio.isocalendar().week


def nombre_archivo_salida(semana, central_norm):
    numero_pms = obtener_numero_pms(semana)
    central_limpia = central_norm.replace(" ", "_")
    return f"PMS_{numero_pms}_PROGRAMA_UNICO_{central_limpia}_{semana}.xlsx"


def generar_programa_unico(
    supabase,
    semana,
    central,
    salida_path=None,
    plantilla_path=None,
):
    """
    Genera el Excel consolidado manteniendo la plantilla.

    Nuevo comportamiento:
    - Usa datos_originales.columnas_excel como fuente principal.
    - Copia C:AU desde cada fila original al Excel consolidado.
    - Mantiene formato de la plantilla destino.
    - Filtra por semana y central.
    """

    semana_inicio = parse_semana_inicio(semana)
    central_norm = normalizar_central(central)

    if central_norm not in ["SANTA ROSA", "VENTANILLA"]:
        raise ValueError("Central no válida. Usa SANTA ROSA o VENTANILLA.")

    if plantilla_path is None:
        plantilla_path = Path(__file__).resolve().parent / PLANTILLA_NOMBRE
    else:
        plantilla_path = Path(plantilla_path)

    if not plantilla_path.exists():
        raise FileNotFoundError(
            f"No se encontró la plantilla: {plantilla_path}. "
            f"Verifica que exista el archivo '{PLANTILLA_NOMBRE}' en el repo."
        )

    archivos_map = consultar_archivos_semana(supabase, semana, central_norm)
    actividades = consultar_actividades(supabase, semana, central_norm, archivos_map)

    wb = load_workbook(plantilla_path)
    ws = seleccionar_hoja(wb, central_norm)

    dejar_solo_hoja_objetivo(wb, ws, central_norm)
    actualizar_encabezado_semana(ws, semana_inicio)

    preparar_filas_destino(ws, len(actividades))

    for idx, actividad in enumerate(actividades):
        fila = FILA_INICIO_DATOS + idx
        archivo_id = actividad.get("pms_archivo_id")
        archivo_info = archivos_map.get(archivo_id, {})
        escribir_fila_desde_original(ws, fila, actividad, archivo_info)

    if not actividades:
        ws[f"{COLS_FALLBACK['actividad']}{FILA_INICIO_DATOS}"] = (
            f"No hay actividades registradas para {CENTRAL_LABEL[central_norm]} en la semana {semana}."
        )
        ws[f"{COLS_FALLBACK['central']}{FILA_INICIO_DATOS}"] = central_norm

    if salida_path is None:
        tmp_dir = tempfile.mkdtemp(prefix="programa_unico_")
        salida_path = Path(tmp_dir) / nombre_archivo_salida(semana, central_norm)
    else:
        salida_path = Path(salida_path)

    wb.save(salida_path)

    numero_pms = obtener_numero_pms(semana)

    return {
        "ok": True,
        "archivo_generado": str(salida_path),
        "nombre_archivo": salida_path.name,
        "total_actividades": len(actividades),
        "central": central_norm,
        "central_label": CENTRAL_LABEL[central_norm],
        "semana": semana,
        "numero_pms": numero_pms,
        "pms_label": f"PMS {numero_pms}",
    }

